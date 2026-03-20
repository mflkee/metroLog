from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from zipfile import ZipFile

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.models.equipment import Repair
from app.schemas.arshin import ArshinSearchResultRead, ArshinVriDetailRead
from app.services.arshin_service import ArshinService
from app.services.user_service import UserService


def bootstrap_admin(db_engine) -> tuple[str, str]:
    testing_session = sessionmaker(bind=db_engine, autoflush=False, autocommit=False, future=True)
    with testing_session() as session:
        UserService(session).ensure_bootstrap_admin()
    return settings.bootstrap_admin_email, settings.bootstrap_admin_password


async def login_user(
    client: AsyncClient,
    *,
    email: str,
    password: str,
) -> dict[str, object]:
    response = await client.post(
        "/api/v1/auth/login",
        json={
            "email": email,
            "password": password,
        },
    )
    assert response.status_code == 200
    return response.json()


@pytest.mark.anyio
async def test_operator_can_create_filter_and_update_registry_entities(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={
            "name": "Лаборатория 1",
            "description": "Основная папка оборудования",
            "sort_order": 10,
        },
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "group_id": None,
            "object_name": "Химическая лаборатория",
            "equipment_type": "SI",
            "name": "Весы лабораторные",
            "modification": "AX-200",
            "serial_number": "SN-001",
            "manufacture_year": 2024,
            "status": "IN_WORK",
            "current_location_manual": "Шкаф 3",
            "si_verification": {
                "vri_id": "1-SI-001",
                "arshin_url": "https://fgis.gost.ru/fundmetrology/cm/results/1-SI-001",
                "org_title": "ФБУ Тест",
                "mit_number": "12345-01",
                "mit_title": "Весы лабораторные",
                "mit_notation": "AX-200",
                "mi_number": "SN-001",
                "result_docnum": "AA 001234",
                "verification_date": "2026-03-01T00:00:00",
                "valid_date": "2027-03-01T00:00:00",
                "raw_payload_json": {"source": "test"},
                "detail_payload_json": {
                    "miInfo": {
                        "singleMI": {
                            "mitypeNumber": "12345-01",
                            "mitypeTitle": "Весы лабораторные",
                            "mitypeType": "AX-200",
                            "manufactureNum": "SN-001",
                            "manufactureYear": 2024,
                            "modification": "AX-200M",
                        }
                    }
                },
            },
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()
    assert equipment["si_verification"]["vri_id"] == "1-SI-001"

    second_equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "Химическая лаборатория",
            "equipment_type": "OTHER",
            "name": "Стол металлический",
            "status": "ARCHIVED",
        },
    )
    assert second_equipment_response.status_code == 201

    folders_list = await client.get("/api/v1/equipment/folders", headers=headers)
    assert folders_list.status_code == 200
    assert len(folders_list.json()) == 1

    groups_list = await client.get(
        f"/api/v1/equipment/groups?folder_id={folder['id']}",
        headers=headers,
    )
    assert groups_list.status_code == 200
    assert len(groups_list.json()) == 0

    equipment_list = await client.get(
        f"/api/v1/equipment?folder_id={folder['id']}",
        headers=headers,
    )
    assert equipment_list.status_code == 200
    assert len(equipment_list.json()) == 2

    search_response = await client.get(
        "/api/v1/equipment?query=AX-200",
        headers=headers,
    )
    assert search_response.status_code == 200
    assert len(search_response.json()) == 1
    assert search_response.json()[0]["id"] == equipment["id"]

    status_response = await client.get(
        "/api/v1/equipment?status=ARCHIVED",
        headers=headers,
    )
    assert status_response.status_code == 200
    assert len(status_response.json()) == 1
    assert status_response.json()[0]["status"] == "ARCHIVED"

    update_response = await client.patch(
        f"/api/v1/equipment/{equipment['id']}",
        headers=headers,
        json={
            "status": "IN_REPAIR",
            "current_location_manual": "Ремонтный участок",
            "modification": None,
        },
    )
    assert update_response.status_code == 200
    updated_equipment = update_response.json()
    assert updated_equipment["status"] == "IN_REPAIR"
    assert updated_equipment["current_location_manual"] == "Ремонтный участок"
    assert updated_equipment["modification"] is None

    detail_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}",
        headers=headers,
    )
    assert detail_response.status_code == 200
    assert detail_response.json()["folder_id"] == folder["id"]
    assert detail_response.json()["si_verification"]["result_docnum"] == "AA 001234"

    delete_equipment_response = await client.delete(
        f"/api/v1/equipment/{equipment['id']}",
        headers=headers,
    )
    assert delete_equipment_response.status_code == 204

    deleted_equipment_detail = await client.get(
        f"/api/v1/equipment/{equipment['id']}",
        headers=headers,
    )
    assert deleted_equipment_detail.status_code == 404

    delete_folder_response = await client.delete(
        f"/api/v1/equipment/folders/{folder['id']}",
        headers=headers,
    )
    assert delete_folder_response.status_code == 204

    folders_after_delete = await client.get("/api/v1/equipment/folders", headers=headers)
    assert folders_after_delete.status_code == 200
    assert folders_after_delete.json() == []

    archived_after_folder_delete = await client.get(
        "/api/v1/equipment?status=ARCHIVED",
        headers=headers,
    )
    assert archived_after_folder_delete.status_code == 200
    assert archived_after_folder_delete.json() == []


@pytest.mark.anyio
async def test_operator_can_delete_equipment_batch(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Массовое удаление"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_ids: list[int] = []
    for index in range(2):
        equipment_response = await client.post(
            "/api/v1/equipment",
            headers=headers,
            json={
                "folder_id": folder["id"],
                "object_name": "Комната подготовки воды",
                "equipment_type": "OTHER",
                "name": f"Прибор #{index + 1}",
                "status": "IN_WORK",
            },
        )
        assert equipment_response.status_code == 201
        equipment_ids.append(equipment_response.json()["id"])

    delete_response = await client.post(
        "/api/v1/equipment/delete-batch",
        headers=headers,
        json={"equipment_ids": equipment_ids},
    )
    assert delete_response.status_code == 204

    equipment_list = await client.get(
        f"/api/v1/equipment?folder_id={folder['id']}",
        headers=headers,
    )
    assert equipment_list.status_code == 200
    assert equipment_list.json() == []


@pytest.mark.anyio
async def test_operator_can_export_filtered_equipment_registry_to_excel(
    client: AsyncClient,
    db_engine,
) -> None:
    from io import BytesIO

    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Экспорт реестра"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    si_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "ХАЛ",
            "equipment_type": "SI",
            "name": "Весы лабораторные",
            "modification": "AX-200",
            "serial_number": "SN-001",
            "manufacture_year": 2024,
            "status": "IN_WORK",
            "current_location_manual": "Шкаф 3",
            "si_verification": {
                "vri_id": "1-SI-EXPORT",
                "arshin_url": "https://fgis.gost.ru/fundmetrology/cm/results/1-SI-EXPORT",
                "org_title": "ФБУ Тест",
                "mit_number": "12345-01",
                "mit_title": "Весы лабораторные",
                "mit_notation": "AX-200",
                "mi_number": "SN-001",
                "result_docnum": "AA 001234",
                "verification_date": "2026-03-01T00:00:00",
                "valid_date": "2027-03-01T00:00:00",
                "raw_payload_json": {"source": "test"},
                "detail_payload_json": {
                    "miInfo": {
                        "singleMI": {
                            "mitypeNumber": "12345-01",
                            "mitypeTitle": "Весы лабораторные",
                            "mitypeType": "AX-200",
                            "manufactureNum": "SN-001",
                            "manufactureYear": 2024,
                            "modification": "AX-200",
                        }
                    }
                },
            },
        },
    )
    assert si_response.status_code == 201

    other_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "ХАЛ",
            "equipment_type": "OTHER",
            "name": "Стол поверочный",
            "status": "ARCHIVED",
            "current_location_manual": "Архив",
        },
    )
    assert other_response.status_code == 201

    export_response = await client.get(
        f"/api/v1/equipment/export/xlsx?folder_id={folder['id']}&equipment_type=SI",
        headers=headers,
    )
    assert export_response.status_code == 200
    assert (
        export_response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(
        filename=BytesIO(export_response.content),
        read_only=True,
        data_only=True,
    )
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    assert rows[0] == (
        "Папка",
        "Категория",
        "Статус",
        "Наименование",
        "Модификация",
        "Серийный номер",
        "Год выпуска",
        "Объект",
        "Локация",
        "Номер свидетельства",
        "Действительно до",
    )
    assert len(rows) == 2
    assert rows[1][0] == "Экспорт реестра"
    assert rows[1][1] == "SI"
    assert rows[1][3] == "Весы лабораторные"
    assert rows[1][9] == "AA 001234"


@pytest.mark.anyio
async def test_operator_can_export_repair_queue_to_excel(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Экспорт ремонтов"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "ХАЛ",
            "equipment_type": "OTHER",
            "name": "Источник питания",
            "status": "IN_WORK",
            "current_location_manual": "Комната ремонта",
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    create_repair_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        data={
            "route_city": "Ленск",
            "route_destination": "Тюмень",
            "sent_to_repair_at": "2026-03-20",
        },
    )
    assert create_repair_response.status_code == 201

    export_response = await client.get(
        "/api/v1/equipment/repairs/export/xlsx?lifecycle_status=active&query=Источник",
        headers=headers,
    )
    assert export_response.status_code == 200
    assert (
        export_response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(
        filename=BytesIO(export_response.content),
        read_only=True,
        data_only=True,
    )
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    assert rows[0] == (
        "Папка",
        "Группа ремонта",
        "Объект",
        "Категория",
        "Наименование",
        "Модификация",
        "Серийный номер",
        "Местонахождение",
        "Откуда",
        "Куда",
        "Отправлено в ремонт",
        "Дедлайн ремонта",
        "Текущий этап",
        "Макс. просрочка, дн.",
        "Свидетельство",
        "Закрыт",
    )
    assert len(rows) == 2
    assert rows[1][0] == "Экспорт ремонтов"
    assert rows[1][3] == "OTHER"
    assert rows[1][4] == "Источник питания"
    assert rows[1][7] == "Комната ремонта"
    assert rows[1][8] == "Ленск"
    assert rows[1][9] == "Тюмень"
    assert rows[1][10] == "20.03.2026"
    assert rows[1][12] == "В ремонте"


@pytest.mark.anyio
async def test_operator_can_export_verification_queue_to_excel(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Экспорт поверок"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "ХАЛ",
            "equipment_type": "SI",
            "name": "Манометр",
            "modification": "DM2005",
            "serial_number": "SN-V-001",
            "status": "IN_WORK",
            "si_verification": {
                "vri_id": "1-SI-VER-EXPORT",
                "arshin_url": "https://fgis.gost.ru/fundmetrology/cm/results/1-SI-VER-EXPORT",
                "org_title": "ФБУ Тест",
                "mit_number": "12345-01",
                "mit_title": "Манометр",
                "mit_notation": "DM2005",
                "mi_number": "SN-V-001",
                "result_docnum": "CERT-V-001",
                "verification_date": "2026-03-01T00:00:00",
                "valid_date": "2027-03-01T00:00:00",
                "raw_payload_json": {"source": "test"},
                "detail_payload_json": {"miInfo": {"singleMI": {"manufactureYear": 2024}}},
            },
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    create_verification_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/verification",
        headers=headers,
        data={
            "route_city": "Ленск",
            "route_destination": "Иркутск",
            "sent_to_verification_at": "2026-03-20",
        },
    )
    assert create_verification_response.status_code == 201

    export_response = await client.get(
        "/api/v1/equipment/verifications/export/xlsx?lifecycle_status=active&query=Манометр",
        headers=headers,
    )
    assert export_response.status_code == 200
    assert (
        export_response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(
        filename=BytesIO(export_response.content),
        read_only=True,
        data_only=True,
    )
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    assert rows[0] == (
        "Папка",
        "Группа поверки",
        "Объект",
        "Наименование",
        "Модификация",
        "Серийный номер",
        "Откуда",
        "Куда",
        "Отправлено в поверку",
        "Состояние",
        "Свидетельство",
        "Действительно до",
        "Закрыта",
    )
    assert len(rows) == 2
    assert rows[1][0] == "Экспорт поверок"
    assert rows[1][2] == "ХАЛ"
    assert rows[1][3] == "Манометр"
    assert rows[1][4] == "DM2005"
    assert rows[1][5] == "SN-V-001"
    assert rows[1][6] == "Ленск"
    assert rows[1][7] == "Иркутск"
    assert rows[1][8] == "20.03.2026"
    assert rows[1][9] == "Отправлено в поверку"
    assert rows[1][10] == "CERT-V-001"
    assert rows[1][11] == "01.03.2027"


@pytest.mark.anyio
async def test_customer_can_view_registry_but_cannot_mutate(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    admin_headers = {"Authorization": f"Bearer {admin['access_token']}"}

    create_customer_response = await client.post(
        "/api/v1/users",
        headers=admin_headers,
        json={
            "first_name": "Customer",
            "last_name": "Viewer",
            "email": "viewer@example.com",
            "role": "CUSTOMER",
            "is_active": True,
        },
    )
    assert create_customer_response.status_code == 201
    customer_password = create_customer_response.json()["temporary_password"]
    customer = await login_user(client, email="viewer@example.com", password=customer_password)
    customer_headers = {"Authorization": f"Bearer {customer['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=admin_headers,
        json={"name": "Доступная папка"},
    )
    assert folder_response.status_code == 201

    list_response = await client.get("/api/v1/equipment/folders", headers=customer_headers)
    assert list_response.status_code == 200
    assert len(list_response.json()) == 1

    forbidden_response = await client.post(
        "/api/v1/equipment/folders",
        headers=customer_headers,
        json={"name": "Запрещенная папка"},
    )
    assert forbidden_response.status_code == 403
    assert forbidden_response.json()["detail"] == "Operator role is required."


@pytest.mark.anyio
async def test_operator_can_search_arshin_and_si_requires_profile(
    client: AsyncClient,
    db_engine,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    async def fake_search_by_certificate(
        self: ArshinService,
        *,
        certificate_number: str,
        year: int | None = None,
    ):
        assert certificate_number == "AA 001234"
        return [
            ArshinSearchResultRead(
                vri_id="1-SI-SEARCH",
                arshin_url="https://fgis.gost.ru/fundmetrology/cm/results/1-SI-SEARCH",
                mit_title="Манометр",
                mit_notation="DM2005",
                mi_modification="серия 5",
                mi_number="SN-7788",
                result_docnum="AA 001234",
                verification_date="2026-03-10T00:00:00",
                valid_date="2027-03-10T00:00:00",
                raw_payload_json={"mocked": True},
            )
        ]

    async def fake_get_vri_detail(self: ArshinService, *, vri_id: str):
        assert vri_id == "1-SI-SEARCH"
        return ArshinVriDetailRead(
            vri_id=vri_id,
            arshin_url="https://fgis.gost.ru/fundmetrology/cm/results/1-SI-SEARCH",
            certificate_number="AA 001234",
            reg_number="60026-15",
            type_designation="201, 523",
            type_name="Скобы с отсчетным устройством",
            serial_number="00024605",
            manufacture_year=2020,
            modification="серии 523",
            organization='АО "АПЗ"',
            verification_date="07.03.2026",
            valid_until="06.03.2027",
            is_usable=True,
            raw_payload_json={"detail": True},
        )

    monkeypatch.setattr(ArshinService, "search_by_certificate", fake_search_by_certificate)
    monkeypatch.setattr(ArshinService, "get_vri_detail", fake_get_vri_detail)

    search_response = await client.post(
        "/api/v1/arshin/search",
        headers=headers,
        json={
            "certificate_number": "AA 001234",
        },
    )
    assert search_response.status_code == 200
    assert search_response.json()[0]["vri_id"] == "1-SI-SEARCH"
    assert search_response.json()[0]["mi_modification"] == "серия 5"

    detail_response = await client.get(
        "/api/v1/arshin/vri/1-SI-SEARCH",
        headers=headers,
    )
    assert detail_response.status_code == 200
    assert detail_response.json()["modification"] == "серии 523"

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "СИ"},
    )
    folder = folder_response.json()

    invalid_create = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "Узел учета",
            "equipment_type": "SI",
            "name": "Манометр",
            "status": "IN_WORK",
        },
    )
    assert invalid_create.status_code == 422
    assert (
        invalid_create.json()["detail"]
        == "SI equipment must be created from an Arshin search result."
    )


@pytest.mark.anyio
async def test_operator_can_refresh_si_card_with_new_certificate(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "СИ обновление"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "ХАЛ",
            "equipment_type": "SI",
            "name": "Скобы",
            "modification": "старое обозначение",
            "serial_number": "OLD-001",
            "manufacture_year": 2018,
            "status": "IN_WORK",
            "si_verification": {
                "vri_id": "vri-old-1",
                "arshin_url": "https://fgis.gost.ru/fundmetrology/cm/results/vri-old-1",
                "org_title": "Старый поверитель",
                "mit_number": "10000-01",
                "mit_title": "Старый тип",
                "mit_notation": "старое обозначение",
                "mi_number": "OLD-001",
                "result_docnum": "OLD-CERT-1",
                "verification_date": "2025-03-01T00:00:00",
                "valid_date": "2026-03-01T00:00:00",
                "raw_payload_json": {"source": "old"},
                "detail_payload_json": {
                    "miInfo": {
                        "singleMI": {
                            "mitypeNumber": "10000-01",
                            "mitypeTitle": "Старый тип",
                            "mitypeType": "обозначение старое",
                            "manufactureNum": "OLD-001",
                            "manufactureYear": 2018,
                            "modification": "старое обозначение",
                        }
                    },
                    "vriInfo": {
                        "organization": "Старый поверитель",
                        "vrfDate": "2025-03-01",
                        "validDate": "2026-03-01",
                        "applicable": {"certNum": "OLD-CERT-1"},
                    },
                },
            },
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    refresh_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/si/refresh",
        headers=headers,
        json={
            "si_verification": {
                "vri_id": "vri-new-1",
                "arshin_url": "https://fgis.gost.ru/fundmetrology/cm/results/vri-new-1",
                "org_title": 'АО "АПЗ"',
                "mit_number": "60026-15",
                "mit_title": "Скобы с отсчетным устройством",
                "mit_notation": "201, 523",
                "mi_number": "00024605",
                "result_docnum": "С-АСГ/07-03-2026/509468383",
                "verification_date": "2026-03-07T00:00:00",
                "valid_date": "2027-03-06T00:00:00",
                "raw_payload_json": {"source": "new-search"},
                "detail_payload_json": {
                    "miInfo": {
                        "singleMI": {
                            "mitypeNumber": "60026-15",
                            "mitypeTitle": "Скобы с отсчетным устройством",
                            "mitypeType": "201, 523",
                            "manufactureNum": "00024605",
                            "manufactureYear": 2020,
                            "modification": "серии 523",
                        }
                    },
                    "vriInfo": {
                        "organization": 'АО "АПЗ"',
                        "vrfDate": "2026-03-07",
                        "validDate": "2027-03-06",
                        "applicable": {"certNum": "С-АСГ/07-03-2026/509468383"},
                    },
                },
            }
        },
    )
    assert refresh_response.status_code == 200
    refreshed = refresh_response.json()
    assert refreshed["name"] == "Скобы с отсчетным устройством"
    assert refreshed["modification"] == "серии 523"
    assert refreshed["serial_number"] == "00024605"
    assert refreshed["manufacture_year"] == 2020
    assert refreshed["si_verification"]["vri_id"] == "vri-new-1"
    assert refreshed["si_verification"]["result_docnum"] == "С-АСГ/07-03-2026/509468383"
    assert refreshed["si_verification"]["mit_notation"] == "201, 523"
    assert (
        refreshed["si_verification"]["detail_payload_json"]["miInfo"]["singleMI"]["modification"]
        == "серии 523"
    )


@pytest.mark.anyio
async def test_operator_can_bulk_import_si_from_excel(
    client: AsyncClient,
    db_engine,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Массовый импорт СИ"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()
    search_calls: dict[str, int | None] = {}

    async def fake_search_by_certificate(
        self: ArshinService,
        *,
        certificate_number: str,
        year: int | None = None,
    ):
        search_calls[certificate_number] = year
        if certificate_number == "CERT-OK-1":
            return [
                ArshinSearchResultRead(
                    vri_id="vri-bulk-1",
                    arshin_url="https://fgis.gost.ru/fundmetrology/cm/results/vri-bulk-1",
                    org_title='АО "АПЗ"',
                    mit_number="60026-15",
                    mit_title="Скобы с отсчетным устройством",
                    mit_notation="201, 523",
                    mi_modification="серии 523",
                    mi_number="00024605",
                    result_docnum="CERT-OK-1",
                    verification_date="2026-03-07T00:00:00",
                    valid_date="2027-03-06T00:00:00",
                    raw_payload_json={"certificate": certificate_number},
                )
            ]
        return []

    async def fake_get_vri_detail(self: ArshinService, *, vri_id: str):
        assert vri_id == "vri-bulk-1"
        return ArshinVriDetailRead(
            vri_id=vri_id,
            arshin_url="https://fgis.gost.ru/fundmetrology/cm/results/vri-bulk-1",
            certificate_number="CERT-OK-1",
            organization='АО "АПЗ"',
            reg_number="60026-15",
            type_designation="201, 523",
            type_name="Скобы с отсчетным устройством",
            serial_number="00024605",
            manufacture_year=2020,
            modification="серии 523",
            verification_date="07.03.2026",
            valid_until="06.03.2027",
            raw_payload_json={
                "miInfo": {
                    "singleMI": {
                        "mitypeNumber": "60026-15",
                        "mitypeTitle": "Скобы с отсчетным устройством",
                        "mitypeType": "201, 523",
                        "manufactureNum": "00024605",
                        "manufactureYear": 2020,
                        "modification": "серии 523",
                    }
                },
                "vriInfo": {
                    "organization": 'АО "АПЗ"',
                    "vrfDate": "2026-03-07",
                    "validDate": "2027-03-06",
                    "applicable": {"certNum": "CERT-OK-1"},
                },
            },
        )

    monkeypatch.setattr(ArshinService, "search_by_certificate", fake_search_by_certificate)
    monkeypatch.setattr(ArshinService, "get_vri_detail", fake_get_vri_detail)

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Служебная строка"
    sheet["B2"] = "Еще заголовок"
    sheet["D3"] = "Не тот столбец"
    sheet["A4"] = "Пояснение"
    sheet["B5"] = "Документ"
    sheet["C5"] = "Наименование"
    sheet["D5"] = "Дата поверки"
    sheet["E5"] = "Примечание"
    sheet["B6"] = "CERT-OK-1"
    sheet["C6"] = "Скобы"
    sheet["D6"] = "12/01/2025"
    sheet["B7"] = "CERT-OK-1"
    sheet["C7"] = "Дубликат"
    sheet["D7"] = "12/01/2025"
    sheet["B8"] = "CERT-MISSING"
    sheet["C8"] = "Не найдено"
    sheet["D8"] = "05.02.2024"

    from io import BytesIO

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)

    response = await client.post(
        "/api/v1/equipment/si/import",
        headers=headers,
        data={
            "folder_id": str(folder["id"]),
            "object_name": "ХАЛ Северный",
            "status_value": "IN_WORK",
            "current_location_manual": "Поступило по Excel",
        },
        files={
            "file": (
                "si-certificates.xlsx",
                stream.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["total_rows"] == 3
    assert payload["created_count"] == 1
    assert payload["skipped_count"] == 1
    assert payload["error_count"] == 1
    assert payload["rows"][0]["row_number"] == 6
    assert payload["rows"][0]["status"] == "created"
    assert payload["rows"][0]["vri_id"] == "vri-bulk-1"
    assert payload["rows"][1]["row_number"] == 7
    assert payload["rows"][1]["status"] == "skipped"
    assert payload["rows"][2]["row_number"] == 8
    assert payload["rows"][2]["status"] == "error"
    assert search_calls == {
        "CERT-OK-1": 2025,
        "CERT-MISSING": 2024,
    }

    equipment_list_response = await client.get(
        f"/api/v1/equipment?folder_id={folder['id']}",
        headers=headers,
    )
    assert equipment_list_response.status_code == 200
    items = equipment_list_response.json()
    assert len(items) == 1
    assert items[0]["equipment_type"] == "SI"
    assert items[0]["object_name"] == "ХАЛ Северный"
    assert items[0]["si_verification"]["result_docnum"] == "CERT-OK-1"


@pytest.mark.anyio
async def test_equipment_detail_returns_404_for_missing_item(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)

    response = await client.get(
        "/api/v1/equipment/9999",
        headers={"Authorization": f"Bearer {admin['access_token']}"},
    )
    assert response.status_code == 404
    assert response.json()["detail"] == "Equipment not found."


@pytest.mark.anyio
async def test_equipment_attachments_can_be_uploaded_listed_and_downloaded(
    client: AsyncClient,
    db_engine,
    tmp_path,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Вложения"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "Лаборатория",
            "equipment_type": "OTHER",
            "name": "Насос",
            "status": "IN_WORK",
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    original_dir = settings.attachment_storage_dir
    settings.attachment_storage_dir = str(tmp_path / "attachments")
    settings.__dict__.pop("attachment_storage_path", None)

    try:
        upload_response = await client.post(
            f"/api/v1/equipment/{equipment['id']}/attachments",
            headers=headers,
            files={"file": ("passport.pdf", b"test attachment payload", "application/pdf")},
        )
        assert upload_response.status_code == 201
        attachment = upload_response.json()
        assert attachment["file_name"] == "passport.pdf"
        assert attachment["uploaded_by_display_name"]
        assert attachment["file_size"] == len(b"test attachment payload")

        list_response = await client.get(
            f"/api/v1/equipment/{equipment['id']}/attachments",
            headers=headers,
        )
        assert list_response.status_code == 200
        attachments = list_response.json()
        assert len(attachments) == 1
        assert attachments[0]["id"] == attachment["id"]

        download_response = await client.get(
            f"/api/v1/equipment/{equipment['id']}/attachments/{attachment['id']}/download",
            headers=headers,
        )
        assert download_response.status_code == 200
        assert download_response.content == b"test attachment payload"
        assert "attachment; filename=\"passport.pdf\"" in (
            download_response.headers.get("content-disposition") or ""
        )

        delete_response = await client.delete(
            f"/api/v1/equipment/{equipment['id']}/attachments/{attachment['id']}",
            headers=headers,
        )
        assert delete_response.status_code == 204

        attachments_after_delete = await client.get(
            f"/api/v1/equipment/{equipment['id']}/attachments",
            headers=headers,
        )
        assert attachments_after_delete.status_code == 200
        assert attachments_after_delete.json() == []

        download_after_delete = await client.get(
            f"/api/v1/equipment/{equipment['id']}/attachments/{attachment['id']}/download",
            headers=headers,
        )
        assert download_after_delete.status_code == 404
    finally:
        settings.attachment_storage_dir = original_dir
        settings.__dict__.pop("attachment_storage_path", None)


@pytest.mark.anyio
async def test_equipment_comments_can_be_created_and_listed(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Комментарии"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "Лаборатория",
            "equipment_type": "OTHER",
            "name": "Компрессор",
            "status": "IN_WORK",
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    create_comment_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/comments",
        headers=headers,
        json={"text": "Прибор осмотрен, внешних повреждений не выявлено."},
    )
    assert create_comment_response.status_code == 201
    comment = create_comment_response.json()
    assert comment["text"] == "Прибор осмотрен, внешних повреждений не выявлено."
    assert comment["author_display_name"]

    list_comments_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}/comments",
        headers=headers,
    )
    assert list_comments_response.status_code == 200
    comments = list_comments_response.json()
    assert len(comments) == 1
    assert comments[0]["id"] == comment["id"]
    assert comments[0]["text"] == comment["text"]

    update_comment_response = await client.patch(
        f"/api/v1/equipment/{equipment['id']}/comments/{comment['id']}",
        headers=headers,
        json={"text": "Комментарий уточнен после повторного осмотра."},
    )
    assert update_comment_response.status_code == 200
    updated_comment = update_comment_response.json()
    assert updated_comment["text"] == "Комментарий уточнен после повторного осмотра."

    delete_comment_response = await client.delete(
        f"/api/v1/equipment/{equipment['id']}/comments/{comment['id']}",
        headers=headers,
    )
    assert delete_comment_response.status_code == 204

    comments_after_delete_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}/comments",
        headers=headers,
    )
    assert comments_after_delete_response.status_code == 200
    assert comments_after_delete_response.json() == []


@pytest.mark.anyio
async def test_operator_can_create_active_repair_with_route_and_dialog_messages(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Ремонтная папка"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "ХАЛ",
            "equipment_type": "OTHER",
            "name": "Источник питания",
            "status": "IN_WORK",
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    create_repair_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        data={
            "route_city": "Тюмень",
            "route_destination": "ПСП ХАЛ Северный",
            "sent_to_repair_at": "2026-03-19",
            "initial_message_text": "Прибор упакован и передан в ремонт.",
        },
        files=[
            (
                "files",
                (
                    "packing-photo.jpg",
                    b"fake-image-payload",
                    "image/jpeg",
                ),
            ),
        ],
    )
    assert create_repair_response.status_code == 201
    repair = create_repair_response.json()
    assert repair["route_city"] == "Тюмень"
    assert repair["route_destination"] == "ПСП ХАЛ Северный"
    assert repair["sent_to_repair_at"] == "2026-03-19"
    assert repair["repair_deadline_at"] == "2026-06-27"

    equipment_detail_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}",
        headers=headers,
    )
    assert equipment_detail_response.status_code == 200
    equipment_detail = equipment_detail_response.json()
    assert equipment_detail["status"] == "IN_REPAIR"
    assert equipment_detail["active_repair"] is not None
    assert equipment_detail["active_repair"]["id"] == repair["id"]
    assert equipment_detail["active_repair"]["route_city"] == "Тюмень"
    assert equipment_detail["active_repair"]["route_destination"] == "ПСП ХАЛ Северный"

    in_repair_list_response = await client.get(
        f"/api/v1/equipment?folder_id={folder['id']}&status=IN_REPAIR",
        headers=headers,
    )
    assert in_repair_list_response.status_code == 200
    in_repair_items = in_repair_list_response.json()
    assert len(in_repair_items) == 1
    assert in_repair_items[0]["id"] == equipment["id"]

    suggestions_response = await client.get(
        f"/api/v1/equipment/folders/{folder['id']}/suggestions",
        headers=headers,
    )
    assert suggestions_response.status_code == 200
    suggestions = suggestions_response.json()
    assert suggestions["object_names"] == ["ХАЛ"]
    assert suggestions["repair_route_cities"] == ["Тюмень"]
    assert suggestions["repair_route_destinations"] == ["ПСП ХАЛ Северный"]

    repair_messages_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}/repair/messages",
        headers=headers,
    )
    assert repair_messages_response.status_code == 200
    repair_messages = repair_messages_response.json()
    assert len(repair_messages) == 1
    assert repair_messages[0]["text"] == "Прибор упакован и передан в ремонт."
    assert len(repair_messages[0]["attachments"]) == 1
    assert repair_messages[0]["attachments"][0]["file_name"] == "packing-photo.jpg"

    add_message_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair/messages",
        headers=headers,
        data={"text": "Чек о перевозке добавлен позже."},
        files=[
            (
                "files",
                (
                    "transport-check.pdf",
                    b"fake-check-payload",
                    "application/pdf",
                ),
            ),
        ],
    )
    assert add_message_response.status_code == 201
    second_message = add_message_response.json()
    assert second_message["text"] == "Чек о перевозке добавлен позже."
    assert len(second_message["attachments"]) == 1
    assert second_message["attachments"][0]["file_name"] == "transport-check.pdf"

    duplicate_repair_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        data={
            "route_city": "Томск",
            "route_destination": "Второй маршрут",
            "sent_to_repair_at": "2026-03-20",
        },
    )
    assert duplicate_repair_response.status_code == 409
    assert (
        duplicate_repair_response.json()["detail"]
        == "Для этого прибора уже есть активный ремонт."
    )


@pytest.mark.anyio
async def test_repair_queue_lists_active_and_archived_repairs(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Очередь ремонтов"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "Комната ремонта",
            "equipment_type": "SI",
            "name": "Калибратор",
            "serial_number": "REP-001",
            "status": "IN_WORK",
            "current_location_manual": "Склад",
            "si_verification": {
                "vri_id": "repair-queue-si-1",
                "arshin_url": "https://fgis.gost.ru/fundmetrology/cm/results/repair-queue-si-1",
                "mit_number": "20000-01",
                "mit_title": "Калибратор",
                "mit_notation": "CAL-100",
                "mi_number": "REP-001",
                "result_docnum": "CERT-REP-001",
                "verification_date": "2026-03-01T00:00:00",
                "valid_date": "2027-03-01T00:00:00",
                "raw_payload_json": {"source": "test"},
                "detail_payload_json": {
                    "miInfo": {
                        "singleMI": {
                            "mitypeNumber": "20000-01",
                            "mitypeTitle": "Калибратор",
                            "mitypeType": "CAL-100",
                            "manufactureNum": "REP-001",
                            "manufactureYear": 2024,
                            "modification": "серия R",
                        }
                    }
                },
            },
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    sent_to_repair_at = date.today() - timedelta(days=120)
    create_repair_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        data={
            "route_city": "Тюмень",
            "route_destination": "Иркутск",
            "sent_to_repair_at": sent_to_repair_at.isoformat(),
        },
    )
    assert create_repair_response.status_code == 201
    repair = create_repair_response.json()

    active_queue_response = await client.get(
        "/api/v1/equipment/repairs?lifecycle_status=active&query=Иркутск",
        headers=headers,
    )
    assert active_queue_response.status_code == 200
    active_items = active_queue_response.json()
    assert len(active_items) == 1
    assert active_items[0]["repair_id"] == repair["id"]
    assert active_items[0]["equipment_id"] == equipment["id"]
    assert active_items[0]["equipment_type"] == "SI"
    assert active_items[0]["route_city"] == "Тюмень"
    assert active_items[0]["route_destination"] == "Иркутск"
    assert active_items[0]["current_stage_label"] == "В ремонте"
    assert active_items[0]["repair_overdue_days"] == 20
    assert active_items[0]["registration_overdue_days"] == 0
    assert active_items[0]["control_overdue_days"] == 0
    assert active_items[0]["payment_overdue_days"] == 0
    assert active_items[0]["max_overdue_days"] == 20
    assert active_items[0]["has_active_verification"] is False
    assert active_items[0]["result_docnum"] == "CERT-REP-001"

    update_repair_response = await client.patch(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        json={
            "arrived_to_destination_at": (sent_to_repair_at + timedelta(days=3)).isoformat(),
            "sent_from_repair_at": (sent_to_repair_at + timedelta(days=110)).isoformat(),
            "sent_from_irkutsk_at": (sent_to_repair_at + timedelta(days=112)).isoformat(),
            "arrived_to_lensk_at": (sent_to_repair_at + timedelta(days=117)).isoformat(),
            "actually_received_at": (sent_to_repair_at + timedelta(days=119)).isoformat(),
            "incoming_control_at": (sent_to_repair_at + timedelta(days=161)).isoformat(),
            "paid_at": (sent_to_repair_at + timedelta(days=233)).isoformat(),
        },
    )
    assert update_repair_response.status_code == 200
    updated_repair = update_repair_response.json()
    assert updated_repair["sent_from_repair_at"] == (
        sent_to_repair_at + timedelta(days=110)
    ).isoformat()
    assert updated_repair["arrived_to_lensk_at"] == (
        sent_to_repair_at + timedelta(days=117)
    ).isoformat()
    assert updated_repair["paid_at"] == (sent_to_repair_at + timedelta(days=233)).isoformat()

    active_queue_after_update_response = await client.get(
        "/api/v1/equipment/repairs?lifecycle_status=active&query=Калибратор",
        headers=headers,
    )
    assert active_queue_after_update_response.status_code == 200
    active_item = active_queue_after_update_response.json()[0]
    assert active_item["sent_from_repair_at"] == (
        sent_to_repair_at + timedelta(days=110)
    ).isoformat()
    assert active_item["registration_deadline_at"] == (
        sent_to_repair_at + timedelta(days=122)
    ).isoformat()
    assert active_item["control_deadline_at"] == (
        sent_to_repair_at + timedelta(days=159)
    ).isoformat()
    assert active_item["payment_deadline_at"] == (
        sent_to_repair_at + timedelta(days=231)
    ).isoformat()
    assert active_item["current_stage_label"] == "Оплата выполнена"
    assert active_item["repair_overdue_days"] == 17
    assert active_item["registration_overdue_days"] == 0
    assert active_item["control_overdue_days"] == 2
    assert active_item["payment_overdue_days"] == 2
    assert active_item["max_overdue_days"] == 17

    testing_session = sessionmaker(bind=db_engine, autoflush=False, autocommit=False, future=True)
    with testing_session() as session:
        repair_row = session.get(Repair, repair["id"])
        assert repair_row is not None
        repair_row.closed_at = date.today()
        session.commit()

    archived_queue_response = await client.get(
        "/api/v1/equipment/repairs?lifecycle_status=archived&query=Калибратор",
        headers=headers,
    )
    assert archived_queue_response.status_code == 200
    archived_items = archived_queue_response.json()
    assert len(archived_items) == 1
    assert archived_items[0]["repair_id"] == repair["id"]
    assert archived_items[0]["closed_at"] == date.today().isoformat()
    assert archived_items[0]["current_stage_label"] == "Ремонт завершен"


@pytest.mark.anyio
async def test_repair_milestones_reject_invalid_stage_order(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Валидация ремонта"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "Лаборатория",
            "equipment_type": "OTHER",
            "name": "Источник питания",
            "serial_number": "R-ORDER-1",
            "status": "IN_WORK",
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    create_repair_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        data={
            "route_city": "Ленск",
            "route_destination": "Иркутск",
            "sent_to_repair_at": "2026-03-20",
        },
    )
    assert create_repair_response.status_code == 201

    invalid_update_response = await client.patch(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        json={
            "arrived_to_lensk_at": "2026-03-22",
            "sent_from_irkutsk_at": "2026-03-21",
        },
    )
    assert invalid_update_response.status_code == 422
    assert (
        invalid_update_response.json()["detail"]
        == "Этап «Отправлено в Ленск» нельзя указать раньше, чем этап «Прибыло в Иркутск»."
    )


@pytest.mark.anyio
async def test_operator_can_create_repair_batch(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Массовый ремонт"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_ids: list[int] = []
    for index in range(2):
        equipment_response = await client.post(
            "/api/v1/equipment",
            headers=headers,
            json={
                "folder_id": folder["id"],
                "object_name": "Комната подготовки воды",
                "equipment_type": "OTHER",
                "name": f"Прибор ремонта #{index + 1}",
                "status": "IN_WORK",
            },
        )
        assert equipment_response.status_code == 201
        equipment_ids.append(equipment_response.json()["id"])

    batch_response = await client.post(
        "/api/v1/equipment/repairs/bulk",
        headers=headers,
        json={
            "equipment_ids": equipment_ids,
            "batch_name": "Комната подготовки воды / ремонт",
            "route_city": "Ленск",
            "route_destination": "Тюмень",
            "sent_to_repair_at": "2026-03-20",
            "initial_message_text": "Партия приборов отправлена в ремонт.",
        },
    )
    assert batch_response.status_code == 201
    payload = batch_response.json()
    assert len(payload) == 2
    assert payload[0]["batch_name"] == "Комната подготовки воды / ремонт"
    assert payload[0]["batch_key"] is not None
    assert payload[1]["batch_key"] == payload[0]["batch_key"]
    assert {item["equipment_id"] for item in payload} == set(equipment_ids)

    repair_queue_response = await client.get(
        "/api/v1/equipment/repairs?lifecycle_status=active&query=Комната подготовки воды / ремонт",
        headers=headers,
    )
    assert repair_queue_response.status_code == 200
    repair_queue = repair_queue_response.json()
    assert len(repair_queue) == 2
    assert all(item["batch_name"] == "Комната подготовки воды / ремонт" for item in repair_queue)
    assert all(item["route_city"] == "Ленск" for item in repair_queue)
    assert all(item["route_destination"] == "Тюмень" for item in repair_queue)


@pytest.mark.anyio
async def test_grouped_repair_shares_messages_updates_and_closes(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Групповой ремонт"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_ids: list[int] = []
    for index in range(2):
        equipment_response = await client.post(
            "/api/v1/equipment",
            headers=headers,
            json={
                "folder_id": folder["id"],
                "object_name": "Подготовка воды",
                "equipment_type": "OTHER",
                "name": f"Ремонтный прибор #{index + 1}",
                "status": "IN_WORK",
            },
        )
        assert equipment_response.status_code == 201
        equipment_ids.append(equipment_response.json()["id"])

    batch_response = await client.post(
        "/api/v1/equipment/repairs/bulk",
        headers=headers,
        json={
            "equipment_ids": equipment_ids,
            "batch_name": "Общая партия ремонта",
            "route_city": "Ленск",
            "route_destination": "Тюмень",
            "sent_to_repair_at": "2026-03-20",
            "initial_message_text": "Ящик с приборами отправлен в ремонт.",
        },
    )
    assert batch_response.status_code == 201
    payload = batch_response.json()
    assert len(payload) == 2
    batch_key = payload[0]["batch_key"]
    assert batch_key is not None
    assert payload[1]["batch_key"] == batch_key

    first_messages_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[0]}/repair/messages",
        headers=headers,
    )
    assert first_messages_response.status_code == 200
    first_messages = first_messages_response.json()
    assert len(first_messages) == 1
    assert first_messages[0]["text"] == "Ящик с приборами отправлен в ремонт."

    second_messages_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[1]}/repair/messages",
        headers=headers,
    )
    assert second_messages_response.status_code == 200
    second_messages = second_messages_response.json()
    assert len(second_messages) == 1
    assert second_messages[0]["text"] == "Ящик с приборами отправлен в ремонт."

    add_group_message_response = await client.post(
        f"/api/v1/equipment/{equipment_ids[1]}/repair/messages",
        headers=headers,
        data={"text": "Группа принята в Тюмени."},
    )
    assert add_group_message_response.status_code == 201

    messages_from_first_after_update = await client.get(
        f"/api/v1/equipment/{equipment_ids[0]}/repair/messages",
        headers=headers,
    )
    assert messages_from_first_after_update.status_code == 200
    shared_messages = messages_from_first_after_update.json()
    assert len(shared_messages) == 2
    assert shared_messages[1]["text"] == "Группа принята в Тюмени."

    update_batch_response = await client.patch(
        f"/api/v1/equipment/repairs/batch/{batch_key}",
        headers=headers,
        json={
            "arrived_to_destination_at": "2026-03-21",
            "sent_from_repair_at": "2026-03-24",
            "sent_from_irkutsk_at": "2026-03-25",
            "arrived_to_lensk_at": "2026-03-27",
            "actually_received_at": "2026-03-28",
            "incoming_control_at": "2026-03-29",
            "paid_at": "2026-03-30",
        },
    )
    assert update_batch_response.status_code == 200
    updated_batch = update_batch_response.json()
    assert len(updated_batch) == 2
    assert all(item["paid_at"] == "2026-03-30" for item in updated_batch)

    close_batch_response = await client.post(
        f"/api/v1/equipment/repairs/batch/{batch_key}/close",
        headers=headers,
    )
    assert close_batch_response.status_code == 200
    closed_batch = close_batch_response.json()
    assert len(closed_batch) == 2
    assert all(item["closed_at"] == date.today().isoformat() for item in closed_batch)

    for equipment_id in equipment_ids:
        equipment_detail_response = await client.get(
            f"/api/v1/equipment/{equipment_id}",
            headers=headers,
        )
        assert equipment_detail_response.status_code == 200
        equipment_detail = equipment_detail_response.json()
        assert equipment_detail["status"] == "IN_WORK"
        assert equipment_detail["active_repair"] is None


@pytest.mark.anyio
async def test_repair_can_be_closed_only_after_payment(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Закрытие ремонта"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "Ремонтная зона",
            "equipment_type": "OTHER",
            "name": "Блок питания",
            "status": "IN_WORK",
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    create_repair_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        data={
            "route_city": "Ленск",
            "route_destination": "Тюмень",
            "sent_to_repair_at": "2026-03-20",
        },
    )
    assert create_repair_response.status_code == 201

    close_without_payment_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair/close",
        headers=headers,
    )
    assert close_without_payment_response.status_code == 422
    assert (
        close_without_payment_response.json()["detail"]
        == "Ремонт можно завершить только после даты оплаты."
    )

    update_repair_response = await client.patch(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        json={
            "arrived_to_destination_at": "2026-03-21",
            "sent_from_repair_at": "2026-03-25",
            "sent_from_irkutsk_at": "2026-03-26",
            "arrived_to_lensk_at": "2026-03-28",
            "actually_received_at": "2026-03-29",
            "incoming_control_at": "2026-03-30",
            "paid_at": "2026-04-01",
        },
    )
    assert update_repair_response.status_code == 200

    close_with_payment_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair/close",
        headers=headers,
    )
    assert close_with_payment_response.status_code == 200
    assert close_with_payment_response.json()["closed_at"] == date.today().isoformat()

    equipment_detail_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}",
        headers=headers,
    )
    assert equipment_detail_response.status_code == 200
    equipment_detail = equipment_detail_response.json()
    assert equipment_detail["status"] == "IN_WORK"
    assert equipment_detail["active_repair"] is None


@pytest.mark.anyio
async def test_closed_repair_is_visible_in_history_and_archive_zip(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Архив ремонтов"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "ХАЛ",
            "equipment_type": "OTHER",
            "name": "Источник питания",
            "status": "IN_WORK",
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    create_repair_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        data={
            "route_city": "Тюмень",
            "route_destination": "Иркутск",
            "sent_to_repair_at": "2026-03-19",
            "initial_message_text": "Прибор упакован и отправлен в ремонт.",
        },
        files=[
            (
                "files",
                (
                    "packing-photo.jpg",
                    b"fake-image-payload",
                    "image/jpeg",
                ),
            ),
        ],
    )
    assert create_repair_response.status_code == 201
    repair = create_repair_response.json()

    add_message_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair/messages",
        headers=headers,
        data={"text": "Добавлен чек по ремонту."},
        files=[
            (
                "files",
                (
                    "repair-check.pdf",
                    b"fake-check-payload",
                    "application/pdf",
                ),
            ),
        ],
    )
    assert add_message_response.status_code == 201

    update_repair_response = await client.patch(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        json={
          "arrived_to_destination_at": "2026-03-21",
          "sent_from_repair_at": "2026-03-25",
          "sent_from_irkutsk_at": "2026-03-26",
          "arrived_to_lensk_at": "2026-03-28",
          "actually_received_at": "2026-03-29",
          "incoming_control_at": "2026-03-30",
          "paid_at": "2026-04-01"
        },
    )
    assert update_repair_response.status_code == 200

    close_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair/close",
        headers=headers,
    )
    assert close_response.status_code == 200
    assert close_response.json()["closed_at"] == date.today().isoformat()

    history_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}/repair/history",
        headers=headers,
    )
    assert history_response.status_code == 200
    history = history_response.json()
    assert len(history) == 1
    assert history[0]["repair_id"] == repair["id"]
    assert history[0]["current_stage_label"] == "Ремонт завершен"
    assert history[0]["closed_at"] == date.today().isoformat()

    archive_response = await client.get(
        f"/api/v1/equipment/repairs/{repair['id']}/archive.zip",
        headers=headers,
    )
    assert archive_response.status_code == 200

    archive = ZipFile(BytesIO(archive_response.content))
    names = archive.namelist()
    assert "dialog.txt" in names
    assert any(name.startswith("files/") and name.endswith("packing-photo.jpg") for name in names)
    assert any(name.startswith("files/") and name.endswith("repair-check.pdf") for name in names)

    dialog_text = archive.read("dialog.txt").decode("utf-8")
    assert "Прибор упакован и отправлен в ремонт." in dialog_text
    assert "Добавлен чек по ремонту." in dialog_text


@pytest.mark.anyio
async def test_si_can_have_independent_active_verification_with_its_own_dialog(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Поверочная папка"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "ХАЛ",
            "equipment_type": "SI",
            "name": "Манометр",
            "status": "IN_WORK",
            "si_verification": {
                "vri_id": "vri-verification-1",
                "arshin_url": "https://fgis.gost.ru/fundmetrology/cm/results/vri-verification-1",
                "mit_number": "10000-01",
                "mit_title": "Манометр",
                "mit_notation": "DM2005",
                "mi_number": "SN-VER-001",
                "result_docnum": "CERT-VER-001",
                "verification_date": "2026-03-01T00:00:00",
                "valid_date": "2027-03-01T00:00:00",
                "raw_payload_json": {"source": "test"},
                "detail_payload_json": {
                    "miInfo": {
                        "singleMI": {
                            "mitypeNumber": "10000-01",
                            "mitypeTitle": "Манометр",
                            "mitypeType": "DM2005",
                            "manufactureNum": "SN-VER-001",
                            "manufactureYear": 2024,
                            "modification": "серия А",
                        }
                    }
                },
            },
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    repair_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        data={
            "route_city": "Тюмень",
            "route_destination": "Ремонтный участок",
            "sent_to_repair_at": "2026-03-19",
        },
    )
    assert repair_response.status_code == 201

    create_verification_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/verification",
        headers=headers,
        data={
            "route_city": "Тюмень",
            "route_destination": "Поверочная лаборатория",
            "sent_to_verification_at": "2026-03-20",
            "initial_message_text": "Прибор подготовлен к поверке.",
        },
        files=[
            (
                "files",
                (
                    "verification-photo.jpg",
                    b"fake-verification-image",
                    "image/jpeg",
                ),
            ),
        ],
    )
    assert create_verification_response.status_code == 201
    verification = create_verification_response.json()
    assert verification["route_city"] == "Тюмень"
    assert verification["route_destination"] == "Поверочная лаборатория"
    assert verification["sent_to_verification_at"] == "2026-03-20"

    verification_messages_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}/verification/messages",
        headers=headers,
    )
    assert verification_messages_response.status_code == 200
    verification_messages = verification_messages_response.json()
    assert len(verification_messages) == 1
    assert verification_messages[0]["text"] == "Прибор подготовлен к поверке."
    assert len(verification_messages[0]["attachments"]) == 1
    assert verification_messages[0]["attachments"][0]["file_name"] == "verification-photo.jpg"

    add_verification_message_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/verification/messages",
        headers=headers,
        data={"text": "Добавлен акт передачи в поверку."},
        files=[
            (
                "files",
                (
                    "verification-act.pdf",
                    b"fake-verification-act",
                    "application/pdf",
                ),
            ),
        ],
    )
    assert add_verification_message_response.status_code == 201
    second_message = add_verification_message_response.json()
    assert second_message["text"] == "Добавлен акт передачи в поверку."
    assert len(second_message["attachments"]) == 1
    assert second_message["attachments"][0]["file_name"] == "verification-act.pdf"

    update_milestones_response = await client.patch(
        f"/api/v1/equipment/{equipment['id']}/verification",
        headers=headers,
        json={
            "received_at_destination_at": "2026-03-21",
            "handed_to_csm_at": "2026-03-22",
        },
    )
    assert update_milestones_response.status_code == 200
    updated_verification = update_milestones_response.json()
    assert updated_verification["received_at_destination_at"] == "2026-03-21"
    assert updated_verification["handed_to_csm_at"] == "2026-03-22"

    equipment_detail_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}",
        headers=headers,
    )
    assert equipment_detail_response.status_code == 200
    equipment_detail = equipment_detail_response.json()
    assert equipment_detail["status"] == "IN_REPAIR"
    assert equipment_detail["active_repair"] is not None
    assert equipment_detail["active_verification"] is not None
    assert equipment_detail["active_verification"]["route_destination"] == "Поверочная лаборатория"
    assert equipment_detail["active_verification"]["received_at_destination_at"] == "2026-03-21"
    assert equipment_detail["active_verification"]["handed_to_csm_at"] == "2026-03-22"

    verification_messages_after_update_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}/verification/messages",
        headers=headers,
    )
    assert verification_messages_after_update_response.status_code == 200
    verification_messages_after_update = verification_messages_after_update_response.json()
    assert len(verification_messages_after_update) == 4
    assert (
        "Получено в пункте назначения (Поверочная лаборатория)"
        in verification_messages_after_update[2]["text"]
    )
    assert "Передано в ЦСМ" in verification_messages_after_update[3]["text"]

    duplicate_verification_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/verification",
        headers=headers,
        data={
            "route_city": "Екатеринбург",
            "route_destination": "Вторая поверка",
            "sent_to_verification_at": "2026-03-21",
        },
    )
    assert duplicate_verification_response.status_code == 409
    assert (
        duplicate_verification_response.json()["detail"]
        == "Для этого прибора уже есть активная поверка."
    )

    close_verification_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/verification/close",
        headers=headers,
    )
    assert close_verification_response.status_code == 200

    equipment_detail_after_close_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}",
        headers=headers,
    )
    assert equipment_detail_after_close_response.status_code == 200
    equipment_detail_after_close = equipment_detail_after_close_response.json()
    assert equipment_detail_after_close["status"] == "IN_REPAIR"
    assert equipment_detail_after_close["active_repair"] is not None
    assert equipment_detail_after_close["active_verification"] is None


@pytest.mark.anyio
async def test_closing_verification_restores_equipment_status_to_in_work(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Закрытие поверки"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "ХАЛ",
            "equipment_type": "SI",
            "name": "Термометр",
            "status": "IN_WORK",
            "si_verification": {
                "vri_id": "vri-close-verification-1",
                "arshin_url": "https://fgis.gost.ru/fundmetrology/cm/results/vri-close-verification-1",
                "mit_number": "30000-01",
                "mit_title": "Термометр",
                "mit_notation": "TM-100",
                "mi_number": "THERM-001",
                "result_docnum": "CERT-CLOSE-001",
                "verification_date": "2026-03-01T00:00:00",
                "valid_date": "2027-03-01T00:00:00",
                "raw_payload_json": {"source": "test"},
                "detail_payload_json": {
                    "miInfo": {
                        "singleMI": {
                            "mitypeNumber": "30000-01",
                            "mitypeTitle": "Термометр",
                            "mitypeType": "TM-100",
                            "manufactureNum": "THERM-001",
                            "manufactureYear": 2024,
                            "modification": "серия T",
                        }
                    }
                },
            },
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    create_verification_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/verification",
        headers=headers,
        data={
            "route_city": "Тюмень",
            "route_destination": "ЦСМ",
            "sent_to_verification_at": "2026-03-20",
        },
    )
    assert create_verification_response.status_code == 201

    close_verification_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/verification/close",
        headers=headers,
    )
    assert close_verification_response.status_code == 200

    equipment_detail_response = await client.get(
        f"/api/v1/equipment/{equipment['id']}",
        headers=headers,
    )
    assert equipment_detail_response.status_code == 200
    equipment_detail = equipment_detail_response.json()
    assert equipment_detail["status"] == "IN_WORK"
    assert equipment_detail["active_repair"] is None
    assert equipment_detail["active_verification"] is None


@pytest.mark.anyio
async def test_verification_queue_lists_active_si_verifications(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Очередь поверки"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "ХАЛ",
            "equipment_type": "SI",
            "name": "Скоба",
            "modification": "серии 523",
            "serial_number": "00024605",
            "manufacture_year": 2020,
            "status": "IN_WORK",
            "si_verification": {
                "vri_id": "vri-queue-1",
                "arshin_url": "https://fgis.gost.ru/fundmetrology/cm/results/vri-queue-1",
                "mit_number": "60026-15",
                "mit_title": "Скобы с отсчетным устройством",
                "mit_notation": "201, 523",
                "mi_number": "00024605",
                "result_docnum": "С-АСГ/07-03-2026/509468383",
                "verification_date": "2026-03-07T00:00:00",
                "valid_date": "2027-03-06T00:00:00",
                "raw_payload_json": {"source": "test"},
                "detail_payload_json": {
                    "miInfo": {
                        "singleMI": {
                            "mitypeNumber": "60026-15",
                            "mitypeTitle": "Скобы с отсчетным устройством",
                            "mitypeType": "201, 523",
                            "manufactureNum": "00024605",
                            "manufactureYear": 2020,
                            "modification": "серии 523",
                        }
                    }
                },
            },
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    repair_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/repair",
        headers=headers,
        data={
            "route_city": "Тюмень",
            "route_destination": "Ремонтный участок",
            "sent_to_repair_at": "2026-03-19",
        },
    )
    assert repair_response.status_code == 201

    verification_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/verification",
        headers=headers,
        data={
            "route_city": "Арзамас",
            "route_destination": "Поверочная лаборатория",
            "sent_to_verification_at": "2026-03-20",
        },
    )
    assert verification_response.status_code == 201

    queue_response = await client.get(
        "/api/v1/equipment/verifications",
        headers=headers,
        params={
            "lifecycle_status": "active",
            "query": "509468383",
        },
    )
    assert queue_response.status_code == 200
    payload = queue_response.json()
    assert len(payload) == 1
    assert payload[0]["equipment_id"] == equipment["id"]
    assert payload[0]["equipment_name"] == "Скоба"
    assert payload[0]["route_city"] == "Арзамас"
    assert payload[0]["route_destination"] == "Поверочная лаборатория"
    assert payload[0]["result_docnum"] == "С-АСГ/07-03-2026/509468383"
    assert payload[0]["has_active_repair"] is True


@pytest.mark.anyio
async def test_verification_milestones_reject_invalid_stage_order(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Валидация поверки"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_response = await client.post(
        "/api/v1/equipment",
        headers=headers,
        json={
            "folder_id": folder["id"],
            "object_name": "Комната СИ",
            "equipment_type": "SI",
            "name": "Манометр",
            "serial_number": "V-ORDER-1",
            "status": "IN_WORK",
            "si_verification": {
                "vri_id": "verification-order-si-1",
                "arshin_url": "https://fgis.gost.ru/fundmetrology/cm/results/verification-order-si-1",
                "mit_number": "20000-99",
                "mit_title": "Манометр",
                "mit_notation": "M-10",
                "mi_number": "V-ORDER-1",
                "result_docnum": "CERT-V-ORDER-1",
                "verification_date": "2026-03-01T00:00:00",
                "valid_date": "2027-03-01T00:00:00",
                "raw_payload_json": {"source": "test"},
                "detail_payload_json": {"miInfo": {"singleMI": {"manufactureYear": 2025}}},
            },
        },
    )
    assert equipment_response.status_code == 201
    equipment = equipment_response.json()

    create_verification_response = await client.post(
        f"/api/v1/equipment/{equipment['id']}/verification",
        headers=headers,
        data={
            "route_city": "Ленск",
            "route_destination": "Иркутск",
            "sent_to_verification_at": "2026-03-20",
        },
    )
    assert create_verification_response.status_code == 201

    invalid_update_response = await client.patch(
        f"/api/v1/equipment/{equipment['id']}/verification",
        headers=headers,
        json={
            "handed_to_csm_at": "2026-03-21",
            "received_at_destination_at": None,
        },
    )
    assert invalid_update_response.status_code == 422
    assert (
        invalid_update_response.json()["detail"]
        == "Этап «Передано в ЦСМ» нельзя указать раньше, чем этап «Получение в Иркутск»."
    )


@pytest.mark.anyio
async def test_bulk_verification_creates_grouped_records(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Групповая поверка"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_ids: list[int] = []
    for index in range(2):
        equipment_response = await client.post(
            "/api/v1/equipment",
            headers=headers,
            json={
                "folder_id": folder["id"],
                "object_name": "Комната подготовки воды",
                "equipment_type": "SI",
                "name": f"Манометр #{index + 1}",
                "status": "IN_WORK",
                "si_verification": {
                    "vri_id": f"bulk-vri-{index + 1}",
                    "arshin_url": (
                        "https://fgis.gost.ru/fundmetrology/cm/results/"
                        f"bulk-vri-{index + 1}"
                    ),
                    "mit_number": "10000-01",
                    "mit_title": "Манометр",
                    "mit_notation": "DM2005",
                    "mi_number": f"SN-BULK-00{index + 1}",
                    "result_docnum": f"CERT-BULK-00{index + 1}",
                    "verification_date": "2026-03-01T00:00:00",
                    "valid_date": "2027-03-01T00:00:00",
                    "raw_payload_json": {"source": "test"},
                    "detail_payload_json": {"miInfo": {"singleMI": {"manufactureYear": 2024}}},
                },
            },
        )
        assert equipment_response.status_code == 201
        equipment_ids.append(equipment_response.json()["id"])

    bulk_response = await client.post(
        "/api/v1/equipment/verifications/bulk",
        headers=headers,
        json={
            "equipment_ids": equipment_ids,
            "batch_name": "Комната подготовки воды / март",
            "route_city": "Ленск",
            "route_destination": "Иркутск",
            "sent_to_verification_at": "2026-03-20",
            "initial_message_text": "Ящик с приборами отправлен в поверку.",
        },
    )
    assert bulk_response.status_code == 201
    payload = bulk_response.json()
    assert len(payload) == 2
    assert payload[0]["batch_name"] == "Комната подготовки воды / март"
    assert payload[0]["batch_key"] is not None
    assert payload[1]["batch_key"] == payload[0]["batch_key"]

    queue_response = await client.get(
        "/api/v1/equipment/verifications",
        headers=headers,
        params={"lifecycle_status": "active", "query": "Комната подготовки воды / март"},
    )
    assert queue_response.status_code == 200
    queue_payload = queue_response.json()
    assert len(queue_payload) == 2
    assert all(row["batch_name"] == "Комната подготовки воды / март" for row in queue_payload)

    first_messages_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[0]}/verification/messages",
        headers=headers,
    )
    assert first_messages_response.status_code == 200
    first_messages = first_messages_response.json()
    assert len(first_messages) == 1
    assert first_messages[0]["text"] == "Ящик с приборами отправлен в поверку."

    second_messages_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[1]}/verification/messages",
        headers=headers,
    )
    assert second_messages_response.status_code == 200
    second_messages = second_messages_response.json()
    assert len(second_messages) == 1
    assert second_messages[0]["text"] == "Ящик с приборами отправлен в поверку."

    add_group_message_response = await client.post(
        f"/api/v1/equipment/{equipment_ids[1]}/verification/messages",
        headers=headers,
        data={"text": "Группа получена в Иркутске."},
    )
    assert add_group_message_response.status_code == 201

    messages_from_first_after_update = await client.get(
        f"/api/v1/equipment/{equipment_ids[0]}/verification/messages",
        headers=headers,
    )
    assert messages_from_first_after_update.status_code == 200
    shared_messages = messages_from_first_after_update.json()
    assert len(shared_messages) == 2
    assert shared_messages[1]["text"] == "Группа получена в Иркутске."

    update_batch_response = await client.patch(
        f"/api/v1/equipment/verifications/batch/{payload[0]['batch_key']}",
        headers=headers,
        json={"received_at_destination_at": "2026-03-21"},
    )
    assert update_batch_response.status_code == 200
    updated_batch = update_batch_response.json()
    assert len(updated_batch) == 2
    assert all(item["received_at_destination_at"] == "2026-03-21" for item in updated_batch)


@pytest.mark.anyio
async def test_operator_can_add_and_remove_repair_batch_items(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Редактирование группы ремонта"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_ids: list[int] = []
    for index in range(3):
        equipment_response = await client.post(
            "/api/v1/equipment",
            headers=headers,
            json={
                "folder_id": folder["id"],
                "object_name": "Подготовка воды",
                "equipment_type": "OTHER",
                "name": f"Партия ремонта #{index + 1}",
                "serial_number": f"R-BATCH-{index + 1}",
                "status": "IN_WORK",
            },
        )
        assert equipment_response.status_code == 201
        equipment_ids.append(equipment_response.json()["id"])

    batch_response = await client.post(
        "/api/v1/equipment/repairs/bulk",
        headers=headers,
        json={
            "equipment_ids": equipment_ids[:2],
            "batch_name": "Партия ремонта / апрель",
            "route_city": "Ленск",
            "route_destination": "Тюмень",
            "sent_to_repair_at": "2026-03-20",
            "initial_message_text": "Партия приборов отправлена в ремонт.",
        },
    )
    assert batch_response.status_code == 201
    batch_key = batch_response.json()[0]["batch_key"]
    assert batch_key is not None

    update_batch_response = await client.patch(
        f"/api/v1/equipment/repairs/batch/{batch_key}",
        headers=headers,
        json={"arrived_to_destination_at": "2026-03-21"},
    )
    assert update_batch_response.status_code == 200

    add_item_response = await client.patch(
        f"/api/v1/equipment/repairs/batch/{batch_key}/items",
        headers=headers,
        json={"add_equipment_ids": [equipment_ids[2]]},
    )
    assert add_item_response.status_code == 200
    added_batch = add_item_response.json()
    assert len(added_batch) == 3
    assert all(item["batch_key"] == batch_key for item in added_batch)
    assert all(item["arrived_to_destination_at"] == "2026-03-21" for item in added_batch)

    third_detail_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[2]}",
        headers=headers,
    )
    assert third_detail_response.status_code == 200
    third_detail = third_detail_response.json()
    assert third_detail["active_repair"]["batch_key"] == batch_key
    assert third_detail["active_repair"]["arrived_to_destination_at"] == "2026-03-21"

    third_messages_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[2]}/repair/messages",
        headers=headers,
    )
    assert third_messages_response.status_code == 200
    third_texts = [message["text"] for message in third_messages_response.json() if message["text"]]
    assert third_texts[0] == "Партия приборов отправлена в ремонт."
    assert any("Партия ремонта #3 (зав. № R-BATCH-3)." in text for text in third_texts)

    remove_item_response = await client.patch(
        f"/api/v1/equipment/repairs/batch/{batch_key}/items",
        headers=headers,
        json={"remove_equipment_ids": [equipment_ids[1]]},
    )
    assert remove_item_response.status_code == 200
    remaining_batch = remove_item_response.json()
    assert len(remaining_batch) == 2
    assert {item["equipment_id"] for item in remaining_batch} == {
        equipment_ids[0],
        equipment_ids[2],
    }

    detached_detail_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[1]}",
        headers=headers,
    )
    assert detached_detail_response.status_code == 200
    detached_detail = detached_detail_response.json()
    assert detached_detail["active_repair"] is not None
    assert detached_detail["active_repair"]["batch_key"] is None
    assert detached_detail["active_repair"]["batch_name"] is None

    detached_messages_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[1]}/repair/messages",
        headers=headers,
    )
    assert detached_messages_response.status_code == 200
    detached_texts = [
        message["text"]
        for message in detached_messages_response.json()
        if message["text"]
    ]
    assert "Партия приборов отправлена в ремонт." in detached_texts
    assert any("Партия ремонта #2 (зав. № R-BATCH-2)." in text for text in detached_texts)
    assert any("Прибор выведен из группы" in text for text in detached_texts)


@pytest.mark.anyio
async def test_operator_can_add_and_remove_verification_batch_items(
    client: AsyncClient,
    db_engine,
) -> None:
    admin_email, admin_password = bootstrap_admin(db_engine)
    admin = await login_user(client, email=admin_email, password=admin_password)
    headers = {"Authorization": f"Bearer {admin['access_token']}"}

    folder_response = await client.post(
        "/api/v1/equipment/folders",
        headers=headers,
        json={"name": "Редактирование группы поверки"},
    )
    assert folder_response.status_code == 201
    folder = folder_response.json()

    equipment_ids: list[int] = []
    for index in range(3):
        equipment_response = await client.post(
            "/api/v1/equipment",
            headers=headers,
            json={
                "folder_id": folder["id"],
                "object_name": "Подготовка воды",
                "equipment_type": "SI",
                "name": f"Партия поверки #{index + 1}",
                "serial_number": f"V-BATCH-{index + 1}",
                "status": "IN_WORK",
                "si_verification": {
                    "vri_id": f"batch-edit-vri-{index + 1}",
                    "arshin_url": (
                        "https://fgis.gost.ru/fundmetrology/cm/results/"
                        f"batch-edit-vri-{index + 1}"
                    ),
                    "mit_number": "10000-01",
                    "mit_title": "Манометр",
                    "mit_notation": "DM2005",
                    "mi_number": f"SN-BATCH-EDIT-{index + 1}",
                    "result_docnum": f"CERT-BATCH-EDIT-{index + 1}",
                    "verification_date": "2026-03-01T00:00:00",
                    "valid_date": "2027-03-01T00:00:00",
                    "raw_payload_json": {"source": "test"},
                    "detail_payload_json": {"miInfo": {"singleMI": {"manufactureYear": 2024}}},
                },
            },
        )
        assert equipment_response.status_code == 201
        equipment_ids.append(equipment_response.json()["id"])

    batch_response = await client.post(
        "/api/v1/equipment/verifications/bulk",
        headers=headers,
        json={
            "equipment_ids": equipment_ids[:2],
            "batch_name": "Партия поверки / апрель",
            "route_city": "Ленск",
            "route_destination": "Иркутск",
            "sent_to_verification_at": "2026-03-20",
            "initial_message_text": "Партия приборов отправлена в поверку.",
        },
    )
    assert batch_response.status_code == 201
    batch_key = batch_response.json()[0]["batch_key"]
    assert batch_key is not None

    update_batch_response = await client.patch(
        f"/api/v1/equipment/verifications/batch/{batch_key}",
        headers=headers,
        json={"received_at_destination_at": "2026-03-21"},
    )
    assert update_batch_response.status_code == 200

    add_item_response = await client.patch(
        f"/api/v1/equipment/verifications/batch/{batch_key}/items",
        headers=headers,
        json={"add_equipment_ids": [equipment_ids[2]]},
    )
    assert add_item_response.status_code == 200
    added_batch = add_item_response.json()
    assert len(added_batch) == 3
    assert all(item["batch_key"] == batch_key for item in added_batch)
    assert all(item["received_at_destination_at"] == "2026-03-21" for item in added_batch)

    third_detail_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[2]}",
        headers=headers,
    )
    assert third_detail_response.status_code == 200
    third_detail = third_detail_response.json()
    assert third_detail["active_verification"]["batch_key"] == batch_key
    assert third_detail["active_verification"]["received_at_destination_at"] == "2026-03-21"

    third_messages_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[2]}/verification/messages",
        headers=headers,
    )
    assert third_messages_response.status_code == 200
    third_texts = [message["text"] for message in third_messages_response.json() if message["text"]]
    assert third_texts[0] == "Партия приборов отправлена в поверку."
    assert any("Партия поверки #3 (зав. № V-BATCH-3)." in text for text in third_texts)

    remove_item_response = await client.patch(
        f"/api/v1/equipment/verifications/batch/{batch_key}/items",
        headers=headers,
        json={"remove_equipment_ids": [equipment_ids[1]]},
    )
    assert remove_item_response.status_code == 200
    remaining_batch = remove_item_response.json()
    assert len(remaining_batch) == 2
    assert {item["equipment_id"] for item in remaining_batch} == {
        equipment_ids[0],
        equipment_ids[2],
    }

    detached_detail_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[1]}",
        headers=headers,
    )
    assert detached_detail_response.status_code == 200
    detached_detail = detached_detail_response.json()
    assert detached_detail["active_verification"] is not None
    assert detached_detail["active_verification"]["batch_key"] is None
    assert detached_detail["active_verification"]["batch_name"] is None

    detached_messages_response = await client.get(
        f"/api/v1/equipment/{equipment_ids[1]}/verification/messages",
        headers=headers,
    )
    assert detached_messages_response.status_code == 200
    detached_texts = [
        message["text"] for message in detached_messages_response.json() if message["text"]
    ]
    assert "Партия приборов отправлена в поверку." in detached_texts
    assert any("Партия поверки #2 (зав. № V-BATCH-2)." in text for text in detached_texts)
    assert any("Прибор выведен из группы" in text for text in detached_texts)
