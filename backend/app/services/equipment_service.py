from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.equipment import (
    Equipment,
    EquipmentAttachment,
    EquipmentComment,
    EquipmentFolder,
    EquipmentGroup,
    EquipmentStatus,
    EquipmentType,
    Repair,
    RepairMessage,
    RepairMessageAttachment,
    SIVerification,
    Verification,
    VerificationMessage,
    VerificationMessageAttachment,
)
from app.models.event import EventCategory, EventLog
from app.models.user import User, UserRole
from app.repositories.equipment_repository import (
    EquipmentAttachmentRepository,
    EquipmentCommentRepository,
    EquipmentFolderRepository,
    EquipmentGroupRepository,
    EquipmentRepository,
    RepairMessageAttachmentRepository,
    RepairMessageRepository,
    RepairRepository,
    SIVerificationRepository,
    VerificationMessageAttachmentRepository,
    VerificationMessageRepository,
    VerificationRepository,
)
from app.repositories.event_repository import EventLogRepository
from app.schemas.arshin import ArshinSearchResultRead, ArshinVriDetailRead
from app.schemas.equipment import (
    EquipmentCommentCreateRequest,
    EquipmentCommentUpdateRequest,
    EquipmentCreateRequest,
    EquipmentFolderCreateRequest,
    EquipmentFolderSuggestionsRead,
    EquipmentFolderUpdateRequest,
    EquipmentSIBulkImportResultRead,
    EquipmentSIBulkImportRowRead,
    EquipmentSIRefreshRequest,
    EquipmentUpdateRequest,
    ProcessBatchMembershipUpdateRequest,
    RepairBulkCreateRequest,
    RepairCreateRequest,
    RepairMessageCreateRequest,
    RepairMilestonesUpdateRequest,
    RepairQueueItemRead,
    VerificationBulkCreateRequest,
    VerificationCreateRequest,
    VerificationMessageCreateRequest,
    VerificationMilestonesUpdateRequest,
    VerificationQueueItemRead,
)
from app.services.arshin_service import ArshinService


class EquipmentService:
    def __init__(self, session: Session) -> None:
        self.session = session
        self.folders = EquipmentFolderRepository(session)
        self.groups = EquipmentGroupRepository(session)
        self.equipment = EquipmentRepository(session)
        self.attachments = EquipmentAttachmentRepository(session)
        self.comments = EquipmentCommentRepository(session)
        self.repairs = RepairRepository(session)
        self.repair_messages = RepairMessageRepository(session)
        self.repair_message_attachments = RepairMessageAttachmentRepository(session)
        self.verifications = VerificationRepository(session)
        self.verification_messages = VerificationMessageRepository(session)
        self.verification_message_attachments = VerificationMessageAttachmentRepository(session)
        self.si_verifications = SIVerificationRepository(session)
        self.events = EventLogRepository(session)

    def list_folders(self) -> list[EquipmentFolder]:
        return self.folders.list_all()

    def create_folder(
        self,
        payload: EquipmentFolderCreateRequest,
        *,
        current_user: User,
    ) -> EquipmentFolder:
        name = _normalize_required_text(payload.name, field_label="Folder name")
        description = _normalize_optional_text(payload.description)
        if self.folders.get_by_name(name) is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A folder with this name already exists.",
            )

        folder = EquipmentFolder(
            name=name,
            description=description,
            sort_order=payload.sort_order,
        )
        self.folders.add(folder)
        self._record_event(
            category=EventCategory.EQUIPMENT,
            action="folder_created",
            user=current_user,
            title=f"Создана папка «{folder.name}»",
            description=_build_nonempty_description(
                [
                    _build_named_detail("Описание", folder.description),
                ]
            ),
            folder_id=folder.id,
            folder_name=folder.name,
        )
        self.session.commit()
        self.session.refresh(folder)
        return folder

    def update_folder(
        self,
        *,
        folder_id: int,
        payload: EquipmentFolderUpdateRequest,
        current_user: User,
    ) -> EquipmentFolder:
        folder = self._get_folder(folder_id)
        changed_fields: list[str] = []

        if "name" in payload.model_fields_set:
            name = _normalize_required_text(payload.name, field_label="Folder name")
            existing = self.folders.get_by_name(name)
            if existing is not None and existing.id != folder.id:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="A folder with this name already exists.",
                )
            folder.name = name
            changed_fields.append("название")

        if "description" in payload.model_fields_set:
            folder.description = _normalize_optional_text(payload.description)
            changed_fields.append("описание")

        if "sort_order" in payload.model_fields_set:
            folder.sort_order = payload.sort_order
            changed_fields.append("порядок")

        if changed_fields:
            self._record_event(
                category=EventCategory.EQUIPMENT,
                action="folder_updated",
                user=current_user,
                title=f"Обновлена папка «{folder.name}»",
                description="Изменено: " + ", ".join(changed_fields) + ".",
                folder_id=folder.id,
                folder_name=folder.name,
            )

        self.session.commit()
        self.session.refresh(folder)
        return folder

    def delete_folder(self, *, folder_id: int, current_user: User) -> None:
        folder = self._get_folder(folder_id)
        self._record_event(
            category=EventCategory.EQUIPMENT,
            action="folder_deleted",
            user=current_user,
            title=f"Удалена папка «{folder.name}»",
            folder_id=folder.id,
            folder_name=folder.name,
        )
        self.equipment.delete_by_folder_id(folder_id=folder.id)
        self.folders.delete(folder)
        self.session.commit()

    def list_groups(self, *, folder_id: int | None = None) -> list[EquipmentGroup]:
        if folder_id is not None:
            self._get_folder(folder_id)
        return self.groups.list_by_folder(folder_id=folder_id)

    def get_folder_suggestions(self, *, folder_id: int) -> EquipmentFolderSuggestionsRead:
        self._get_folder(folder_id)
        process_batch_names = sorted(
            {
                *self.repairs.list_distinct_batch_names(folder_id=folder_id),
                *self.verifications.list_distinct_batch_names(folder_id=folder_id),
            }
        )
        return EquipmentFolderSuggestionsRead(
            object_names=self.equipment.list_distinct_object_names(folder_id=folder_id),
            current_locations=self.equipment.list_distinct_locations(folder_id=folder_id),
            repair_route_cities=self.repairs.list_distinct_route_cities(folder_id=folder_id),
            repair_route_destinations=self.repairs.list_distinct_route_destinations(
                folder_id=folder_id
            ),
            process_batch_names=process_batch_names,
        )

    def list_equipment(
        self,
        *,
        folder_id: int | None = None,
        group_id: int | None = None,
        query: str | None = None,
        status: EquipmentStatus | None = None,
        equipment_type: EquipmentType | None = None,
    ) -> list[Equipment]:
        if folder_id is not None:
            self._get_folder(folder_id)
        if group_id is not None:
            self._get_group(group_id)
        return self.equipment.list_all(
            folder_id=folder_id,
            group_id=group_id,
            query=query.strip() if query else None,
            status=status,
            equipment_type=equipment_type,
        )

    def list_verification_queue(
        self,
        *,
        lifecycle_status: str,
        query: str | None = None,
    ) -> list[VerificationQueueItemRead]:
        normalized_lifecycle_status = lifecycle_status.strip().lower()
        if normalized_lifecycle_status not in {"active", "archived"}:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Verification lifecycle status must be active or archived.",
            )

        rows = self.verifications.list_queue_items(
            lifecycle_status=normalized_lifecycle_status,
            query=query.strip() if query else None,
        )
        return [
            self._build_verification_queue_item(
                verification=verification,
                equipment=equipment,
                si_verification=si_verification,
                has_active_repair=has_active_repair,
            )
            for verification, equipment, si_verification, has_active_repair in rows
        ]

    def list_repair_queue(
        self,
        *,
        lifecycle_status: str,
        query: str | None = None,
    ) -> list[RepairQueueItemRead]:
        normalized_lifecycle_status = lifecycle_status.strip().lower()
        if normalized_lifecycle_status not in {"active", "archived"}:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Repair lifecycle status must be active or archived.",
            )

        rows = self.repairs.list_queue_items(
            lifecycle_status=normalized_lifecycle_status,
            query=query.strip() if query else None,
        )
        return [
            self._build_repair_queue_item(
                repair=repair,
                equipment=equipment,
                si_verification=si_verification,
                has_active_verification=has_active_verification,
            )
            for repair, equipment, si_verification, has_active_verification in rows
        ]

    def list_equipment_verification_history(
        self,
        *,
        equipment_id: int,
    ) -> list[VerificationQueueItemRead]:
        self.get_equipment(equipment_id=equipment_id)
        rows = self.verifications.list_archived_by_equipment_id(equipment_id=equipment_id)
        return [
            self._build_verification_queue_item(
                verification=verification,
                equipment=equipment,
                si_verification=si_verification,
                has_active_repair=has_active_repair,
            )
            for verification, equipment, si_verification, has_active_repair in rows
        ]

    def list_equipment_repair_history(
        self,
        *,
        equipment_id: int,
    ) -> list[RepairQueueItemRead]:
        self.get_equipment(equipment_id=equipment_id)
        rows = self.repairs.list_archived_by_equipment_id(equipment_id=equipment_id)
        return [
            self._build_repair_queue_item(
                repair=repair,
                equipment=equipment,
                si_verification=si_verification,
                has_active_verification=has_active_verification,
            )
            for repair, equipment, si_verification, has_active_verification in rows
        ]

    def get_equipment(self, *, equipment_id: int) -> Equipment:
        equipment = self.equipment.get_by_id(equipment_id)
        if equipment is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Equipment not found.",
            )
        return equipment

    def create_equipment(
        self,
        payload: EquipmentCreateRequest,
        *,
        current_user: User | None = None,
    ) -> Equipment:
        folder = self._get_folder(payload.folder_id)
        group = self._get_group(payload.group_id) if payload.group_id is not None else None
        if group is not None and group.folder_id != folder.id:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail="Group does not belong to the selected folder.",
            )
        self._validate_si_payload_for_create(payload)

        equipment = Equipment(
            folder_id=folder.id,
            group_id=payload.group_id,
            object_name=_normalize_required_text(payload.object_name, field_label="Object name"),
            equipment_type=payload.equipment_type,
            name=_normalize_required_text(payload.name, field_label="Equipment name"),
            modification=_normalize_optional_text(payload.modification),
            serial_number=_normalize_optional_text(payload.serial_number),
            manufacture_year=_validate_manufacture_year(payload.manufacture_year),
            status=payload.status,
            current_location_manual=_normalize_optional_text(payload.current_location_manual),
        )
        self.equipment.add(equipment)
        if payload.si_verification is not None:
            self.si_verifications.add(
                SIVerification(
                    equipment_id=equipment.id,
                    vri_id=_normalize_required_text(
                        payload.si_verification.vri_id,
                        field_label="SI vri_id",
                    ),
                    arshin_url=_normalize_long_optional_text(payload.si_verification.arshin_url),
                    org_title=_normalize_optional_text(payload.si_verification.org_title),
                    mit_number=_normalize_optional_text(payload.si_verification.mit_number),
                    mit_title=_normalize_optional_text(payload.si_verification.mit_title),
                    mit_notation=_normalize_optional_text(payload.si_verification.mit_notation),
                    mi_number=_normalize_optional_text(payload.si_verification.mi_number),
                    result_docnum=_normalize_optional_text(payload.si_verification.result_docnum),
                    verification_date=payload.si_verification.verification_date,
                    valid_date=payload.si_verification.valid_date,
                    raw_payload_json=payload.si_verification.raw_payload_json,
                    detail_payload_json=payload.si_verification.detail_payload_json,
                )
            )
        if current_user is not None:
            self._record_equipment_event(
                action="equipment_created",
                user=current_user,
                equipment=equipment,
                title=f"Создан прибор «{equipment.name}»",
                description=_build_nonempty_description(
                    [
                        _build_named_detail("Объект", equipment.object_name),
                        _build_named_detail("Категория", equipment.equipment_type.value),
                        _build_named_detail("Серийный номер", equipment.serial_number),
                    ]
                ),
            )
        self.session.commit()
        self.session.refresh(equipment)
        return equipment

    def update_equipment(
        self,
        *,
        equipment_id: int,
        payload: EquipmentUpdateRequest,
        current_user: User | None = None,
    ) -> Equipment:
        equipment = self.get_equipment(equipment_id=equipment_id)
        next_folder_id = equipment.folder_id
        next_group_id = equipment.group_id
        changed_fields: list[str] = []

        if "folder_id" in payload.model_fields_set:
            if payload.folder_id is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Equipment folder_id must not be empty.",
                )
            self._get_folder(payload.folder_id)
            next_folder_id = payload.folder_id

        if "group_id" in payload.model_fields_set:
            if payload.group_id is not None:
                self._get_group(payload.group_id)
            next_group_id = payload.group_id

        if next_group_id is not None:
            next_group = self._get_group(next_group_id)
            if next_folder_id is None:
                next_folder_id = next_group.folder_id
            if next_group.folder_id != next_folder_id:
                folder_changed = "folder_id" in payload.model_fields_set
                group_changed = "group_id" in payload.model_fields_set
                if folder_changed and not group_changed:
                    next_group_id = None
                else:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="Group does not belong to the selected folder.",
                    )

        equipment.folder_id = next_folder_id
        equipment.group_id = next_group_id
        if {"folder_id", "group_id"} & payload.model_fields_set:
            changed_fields.append("папка/группа")

        if "object_name" in payload.model_fields_set:
            equipment.object_name = _normalize_required_text(
                payload.object_name,
                field_label="Object name",
            )
            changed_fields.append("объект")

        if "equipment_type" in payload.model_fields_set:
            if payload.equipment_type != equipment.equipment_type:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Equipment type cannot be changed after creation.",
                )
            equipment.equipment_type = payload.equipment_type

        if "name" in payload.model_fields_set:
            equipment.name = _normalize_required_text(payload.name, field_label="Equipment name")
            changed_fields.append("наименование")

        if "modification" in payload.model_fields_set:
            equipment.modification = _normalize_optional_text(payload.modification)
            changed_fields.append("модификация")

        if "serial_number" in payload.model_fields_set:
            equipment.serial_number = _normalize_optional_text(payload.serial_number)
            changed_fields.append("заводской номер")

        if "manufacture_year" in payload.model_fields_set:
            equipment.manufacture_year = _validate_manufacture_year(payload.manufacture_year)
            changed_fields.append("год выпуска")

        if "status" in payload.model_fields_set:
            equipment.status = payload.status
            changed_fields.append("статус")

        if "current_location_manual" in payload.model_fields_set:
            equipment.current_location_manual = _normalize_optional_text(
                payload.current_location_manual
            )
            changed_fields.append("местонахождение")

        if current_user is not None and changed_fields:
            self._record_equipment_event(
                action="equipment_updated",
                user=current_user,
                equipment=equipment,
                title=f"Обновлен прибор «{equipment.name}»",
                description="Изменено: " + ", ".join(changed_fields) + ".",
            )

        self.session.commit()
        self.session.refresh(equipment)
        return equipment

    def refresh_si_verification(
        self,
        *,
        equipment_id: int,
        payload: EquipmentSIRefreshRequest,
        current_user: User | None = None,
    ) -> Equipment:
        equipment = self.get_equipment(equipment_id=equipment_id)
        if equipment.equipment_type != EquipmentType.SI:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Only SI equipment supports Arshin refresh.",
            )

        si_payload = payload.si_verification
        if si_payload.detail_payload_json is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="SI refresh requires Arshin detail payload.",
            )

        existing = self.si_verifications.get_by_vri_id(vri_id=si_payload.vri_id.strip())
        if existing is not None and existing.equipment_id != equipment.id:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=self._build_existing_si_message(
                    existing,
                    prefix="Запись Аршина с этим свидетельством уже привязана",
                ),
            )

        detail = si_payload.detail_payload_json
        detail_mi = _extract_si_detail_single_mi(detail)
        detail_vri = _extract_si_detail_vri_info(detail)

        equipment.name = _normalize_required_text(
            _first_nonempty_str(detail_mi.get("mitypeTitle"), si_payload.mit_title, equipment.name),
            field_label="Equipment name",
        )
        equipment.modification = _normalize_optional_text(
            _first_nonempty_str(detail_mi.get("modification"), equipment.modification)
        )
        equipment.serial_number = _normalize_optional_text(
            _first_nonempty_str(
                detail_mi.get("manufactureNum"),
                si_payload.mi_number,
                equipment.serial_number,
            )
        )
        equipment.manufacture_year = _validate_manufacture_year(
            _first_nonempty_int(detail_mi.get("manufactureYear"), equipment.manufacture_year)
        )

        si_verification = self.si_verifications.get_by_equipment_id(equipment_id=equipment.id)
        if si_verification is None:
            si_verification = SIVerification(equipment_id=equipment.id, vri_id="")
            self.si_verifications.add(si_verification)

        si_verification.vri_id = _normalize_required_text(
            si_payload.vri_id,
            field_label="SI vri_id",
        )
        si_verification.arshin_url = _normalize_long_optional_text(si_payload.arshin_url)
        si_verification.org_title = _normalize_optional_text(
            _first_nonempty_str(detail_vri.get("organization"), si_payload.org_title)
        )
        si_verification.mit_number = _normalize_optional_text(
            _first_nonempty_str(detail_mi.get("mitypeNumber"), si_payload.mit_number)
        )
        si_verification.mit_title = _normalize_optional_text(
            _first_nonempty_str(detail_mi.get("mitypeTitle"), si_payload.mit_title)
        )
        si_verification.mit_notation = _normalize_optional_text(
            _first_nonempty_str(detail_mi.get("mitypeType"), si_payload.mit_notation)
        )
        si_verification.mi_number = _normalize_optional_text(
            _first_nonempty_str(detail_mi.get("manufactureNum"), si_payload.mi_number)
        )
        si_verification.result_docnum = _normalize_optional_text(
            _first_nonempty_str(
                _extract_cert_num_from_detail(detail_vri),
                si_payload.result_docnum,
            )
        )
        si_verification.verification_date = _first_nonempty_datetime(
            _parse_date_to_datetime(detail_vri.get("vrfDate")),
            si_payload.verification_date,
        )
        si_verification.valid_date = _first_nonempty_datetime(
            _parse_date_to_datetime(detail_vri.get("validDate")),
            si_payload.valid_date,
        )
        si_verification.raw_payload_json = si_payload.raw_payload_json
        si_verification.detail_payload_json = si_payload.detail_payload_json

        if current_user is not None:
            self._record_equipment_event(
                action="si_refreshed",
                user=current_user,
                equipment=equipment,
                title=f"Обновлены данные СИ «{equipment.name}» из Аршина",
                description=_build_nonempty_description(
                    [
                        _build_named_detail("Свидетельство", si_verification.result_docnum),
                        _build_named_detail(
                            "Действительно до",
                            self._format_sheet_date(si_verification.valid_date),
                        ),
                    ]
                ),
            )

        self.session.commit()
        self.session.refresh(equipment)
        return equipment

    async def import_si_from_excel(
        self,
        *,
        file_name: str | None,
        content: bytes,
        folder_id: int,
        object_name: str,
        status_value: EquipmentStatus,
        current_location_manual: str | None,
        arshin_service: ArshinService | None = None,
        current_user: User | None = None,
    ) -> EquipmentSIBulkImportResultRead:
        self._get_folder(folder_id)

        rows = _extract_certificate_rows_from_table(
            file_name=file_name,
            content=content,
        )
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Excel file does not contain certificate numbers.",
            )

        arshin = arshin_service or ArshinService()
        result_rows: list[EquipmentSIBulkImportRowRead] = []
        seen_certificates: set[str] = set()

        for row in rows:
            row_number = row.row_number
            certificate_number = row.certificate_number
            normalized_certificate = _normalize_certificate_number(certificate_number)
            if normalized_certificate in seen_certificates:
                result_rows.append(
                    EquipmentSIBulkImportRowRead(
                        row_number=row_number,
                        certificate_number=certificate_number,
                        status="skipped",
                        message="Duplicate certificate number inside the uploaded file.",
                    )
                )
                continue
            seen_certificates.add(normalized_certificate)

            try:
                search_results = await arshin.search_by_certificate(
                    certificate_number=certificate_number,
                    year=row.verification_year,
                )
                matched_result = _select_bulk_import_candidate(
                    certificate_number=certificate_number,
                    results=search_results,
                )
                if matched_result is None:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="Arshin search returned multiple ambiguous records.",
                    )

                detail = await arshin.get_vri_detail(vri_id=matched_result.vri_id)
                equipment = self.create_equipment(
                    _build_si_create_request_from_arshin(
                        folder_id=folder_id,
                        object_name=object_name,
                        status_value=status_value,
                        current_location_manual=current_location_manual,
                        result=matched_result,
                        detail=detail,
                    ),
                    current_user=None,
                )
                result_rows.append(
                    EquipmentSIBulkImportRowRead(
                        row_number=row_number,
                        certificate_number=certificate_number,
                        status="created",
                        message="SI equipment item was created from Arshin.",
                        equipment_id=equipment.id,
                        equipment_name=equipment.name,
                        vri_id=matched_result.vri_id,
                    )
                )
            except HTTPException as exc:
                row_status = "skipped" if exc.status_code == status.HTTP_409_CONFLICT else "error"
                result_rows.append(
                    EquipmentSIBulkImportRowRead(
                        row_number=row_number,
                        certificate_number=certificate_number,
                        status=row_status,
                        message=str(exc.detail),
                    )
                )

        result = EquipmentSIBulkImportResultRead(
            total_rows=len(rows),
            created_count=sum(1 for row in result_rows if row.status == "created"),
            skipped_count=sum(1 for row in result_rows if row.status == "skipped"),
            error_count=sum(1 for row in result_rows if row.status == "error"),
            rows=result_rows,
        )
        if current_user is not None:
            folder = self._get_folder(folder_id)
            self._record_event(
                category=EventCategory.EQUIPMENT,
                action="si_imported",
                user=current_user,
                title=f"Выполнен импорт СИ в папку «{folder.name}»",
                description=(
                    f"Создано: {result.created_count}. "
                    f"Пропущено: {result.skipped_count}. "
                    f"Ошибок: {result.error_count}."
                ),
                folder_id=folder.id,
                folder_name=folder.name,
            )
        return result

    def export_equipment_registry_xlsx(
        self,
        *,
        folder_id: int | None = None,
        group_id: int | None = None,
        query: str | None = None,
        status_value: EquipmentStatus | None = None,
        equipment_type: EquipmentType | None = None,
    ) -> bytes:
        equipment_items = self.list_equipment(
            folder_id=folder_id,
            group_id=group_id,
            query=query,
            status=status_value,
            equipment_type=equipment_type,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Equipment"
        headers = [
            "Папка",
            "Категория",
            "Статус",
            "Наименование",
            "Модификация",
            "Серийный номер",
            "Год выпуска",
            "Объект",
            "Локация",
            "Номер свидетельства",
            "Действительно до",
        ]
        sheet.append(headers)

        for equipment in equipment_items:
            folder_name = None
            if equipment.folder_id is not None:
                folder = self.folders.get_by_id(equipment.folder_id)
                folder_name = folder.name if folder is not None else None

            verification = equipment.si_verification
            sheet.append(
                [
                    folder_name,
                    equipment.equipment_type.value,
                    equipment.status.value,
                    equipment.name,
                    equipment.modification,
                    equipment.serial_number,
                    equipment.manufacture_year,
                    equipment.object_name,
                    equipment.current_location_manual,
                    verification.result_docnum if verification is not None else None,
                    (
                        verification.valid_date.strftime("%d.%m.%Y")
                        if verification is not None and verification.valid_date is not None
                        else None
                    ),
                ]
            )

        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def export_repair_queue_xlsx(
        self,
        *,
        lifecycle_status: str,
        query: str | None = None,
    ) -> bytes:
        items = self.list_repair_queue(
            lifecycle_status=lifecycle_status,
            query=query,
        )
        headers = [
            "Папка",
            "Группа ремонта",
            "Объект",
            "Категория",
            "Наименование",
            "Модификация",
            "Серийный номер",
            "Местонахождение",
            "Откуда",
            "Куда",
            "Отправлено в ремонт",
            "Дедлайн ремонта",
            "Текущий этап",
            "Макс. просрочка, дн.",
            "Свидетельство",
            "Закрыт",
        ]
        rows: list[list[object | None]] = []
        for item in items:
            folder_name = None
            if item.folder_id is not None:
                folder = self.folders.get_by_id(item.folder_id)
                folder_name = folder.name if folder is not None else None
            rows.append(
                [
                    folder_name,
                    item.batch_name,
                    item.object_name,
                    item.equipment_type.value,
                    item.equipment_name,
                    item.modification,
                    item.serial_number,
                    item.current_location_manual,
                    item.route_city,
                    item.route_destination,
                    self._format_sheet_date(item.sent_to_repair_at),
                    self._format_sheet_date(item.repair_deadline_at),
                    item.current_stage_label,
                    item.max_overdue_days,
                    item.result_docnum,
                    self._format_sheet_date(item.closed_at),
                ]
            )
        return self._build_workbook_bytes(
            sheet_title="Repairs",
            headers=headers,
            rows=rows,
        )

    def export_verification_queue_xlsx(
        self,
        *,
        lifecycle_status: str,
        query: str | None = None,
    ) -> bytes:
        items = self.list_verification_queue(
            lifecycle_status=lifecycle_status,
            query=query,
        )
        headers = [
            "Папка",
            "Группа поверки",
            "Объект",
            "Наименование",
            "Модификация",
            "Серийный номер",
            "Откуда",
            "Куда",
            "Отправлено в поверку",
            "Состояние",
            "Свидетельство",
            "Действительно до",
            "Закрыта",
        ]
        rows: list[list[object | None]] = []
        for item in items:
            folder_name = None
            if item.folder_id is not None:
                folder = self.folders.get_by_id(item.folder_id)
                folder_name = folder.name if folder is not None else None
            rows.append(
                [
                    folder_name,
                    item.batch_name,
                    item.object_name,
                    item.equipment_name,
                    item.modification,
                    item.serial_number,
                    item.route_city,
                    item.route_destination,
                    self._format_sheet_date(item.sent_to_verification_at),
                    _get_verification_progress_label(item),
                    item.result_docnum,
                    self._format_sheet_date(item.valid_date),
                    self._format_sheet_date(item.closed_at),
                ]
            )
        return self._build_workbook_bytes(
            sheet_title="Verification",
            headers=headers,
            rows=rows,
        )

    def _build_workbook_bytes(
        self,
        *,
        sheet_title: str,
        headers: list[str],
        rows: list[list[object | None]],
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_title
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def _format_sheet_date(self, value: date | datetime | None) -> str | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            value = value.date()
        return value.strftime("%d.%m.%Y")

    def create_repair(
        self,
        *,
        equipment_id: int,
        payload: RepairCreateRequest,
        current_user: User,
        files: list[UploadedFilePayload] | None = None,
    ) -> Repair:
        equipment = self.get_equipment(equipment_id=equipment_id)
        existing_repair = self.repairs.get_active_by_equipment_id(equipment_id=equipment.id)
        if existing_repair is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Для этого прибора уже есть активный ремонт.",
            )

        sent_to_repair_at = payload.sent_to_repair_at
        repair = Repair(
            equipment_id=equipment.id,
            batch_key=_normalize_optional_text(payload.batch_key),
            batch_name=_normalize_optional_text(payload.batch_name),
            route_city=_normalize_required_text(
                payload.route_city,
                field_label="Repair route city",
            ),
            route_destination=_normalize_required_text(
                payload.route_destination,
                field_label="Repair route destination",
            ),
            sent_to_repair_at=sent_to_repair_at,
            repair_deadline_at=sent_to_repair_at + timedelta(days=100),
        )
        equipment.status = EquipmentStatus.IN_REPAIR
        self.repairs.add(repair)
        normalized_initial_text = _normalize_message_text(payload.initial_message_text)
        if normalized_initial_text is not None or files:
            self._create_repair_message_record(
                repair=repair,
                author=current_user,
                payload=RepairMessageCreateRequest(text=normalized_initial_text),
                files=files or [],
            )
        self._record_event(
            category=EventCategory.REPAIR,
            action="repair_created",
            user=current_user,
            equipment=equipment,
            batch_key=repair.batch_key,
            title=f"Прибор «{equipment.name}» отправлен в ремонт",
            description=_build_nonempty_description(
                [
                    _build_named_detail("Откуда", repair.route_city),
                    _build_named_detail("Куда", repair.route_destination),
                    _build_named_detail(
                        "Отправлено",
                        self._format_sheet_date(repair.sent_to_repair_at),
                    ),
                    _build_named_detail("Группа", repair.batch_name),
                ]
            ),
        )
        self.session.commit()
        self.session.refresh(repair)
        return repair

    def create_repair_batch(
        self,
        *,
        payload: RepairBulkCreateRequest,
        current_user: User,
    ) -> list[Repair]:
        equipment_ids = list(dict.fromkeys(payload.equipment_ids))
        if not equipment_ids:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Нужно выбрать хотя бы один прибор для ремонта.",
            )

        batch_key = uuid4().hex
        batch_name = _normalize_required_text(
            payload.batch_name,
            field_label="Repair batch name",
        )
        created: list[Repair] = []
        for index, equipment_id in enumerate(equipment_ids):
            created.append(
                self.create_repair(
                    equipment_id=equipment_id,
                    payload=RepairCreateRequest(
                        batch_key=batch_key,
                        batch_name=batch_name,
                        route_city=payload.route_city,
                        route_destination=payload.route_destination,
                        sent_to_repair_at=payload.sent_to_repair_at,
                        initial_message_text=payload.initial_message_text if index == 0 else None,
                    ),
                    current_user=current_user,
                    files=[],
                )
            )
        return created

    def update_repair_batch_items(
        self,
        *,
        batch_key: str,
        payload: ProcessBatchMembershipUpdateRequest,
        current_user: User,
    ) -> list[Repair]:
        normalized_batch_key = _normalize_required_text(
            batch_key,
            field_label="Repair batch key",
        )
        repairs = self.repairs.list_active_by_batch_key(batch_key=normalized_batch_key)
        if not repairs:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Активная группа ремонта не найдена.",
            )

        anchor = repairs[0]
        existing_ids = {repair.equipment_id for repair in repairs}
        add_ids = [
            equipment_id
            for equipment_id in dict.fromkeys(payload.add_equipment_ids)
            if equipment_id not in existing_ids
        ]
        remove_ids = [
            equipment_id
            for equipment_id in dict.fromkeys(payload.remove_equipment_ids)
            if equipment_id in existing_ids
        ]

        if not add_ids and not remove_ids:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Нужно добавить или удалить хотя бы один прибор из группы ремонта.",
            )
        if len(remove_ids) >= len(repairs):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Нельзя удалить из группы ремонта все приборы сразу.",
            )

        added_labels: list[str] = []
        removed_labels: list[str] = []

        for equipment_id in add_ids:
            new_repair = self._create_repair_from_batch_anchor(
                anchor=anchor,
                equipment_id=equipment_id,
            )
            added_labels.append(_build_equipment_batch_member_label(new_repair.equipment))

        if added_labels:
            self._create_repair_message_record(
                repair=anchor,
                author=current_user,
                payload=RepairMessageCreateRequest(
                    text="В группу ремонта добавлены приборы: "
                    + ", ".join(sorted(added_labels))
                    + "."
                ),
                files=[],
            )
            self._record_event(
                category=EventCategory.REPAIR,
                action="repair_batch_items_added",
                user=current_user,
                title=(
                    "В группу ремонта "
                    f"«{anchor.batch_name or anchor.route_destination}» добавлены приборы"
                ),
                description=f"Добавлено приборов: {len(added_labels)}.",
                equipment=anchor.equipment,
                batch_key=anchor.batch_key,
            )

        message_repair = next(
            (item for item in repairs if item.equipment_id not in remove_ids),
            anchor,
        )
        for repair in repairs:
            if repair.equipment_id not in remove_ids:
                continue
            removed_labels.append(_build_equipment_batch_member_label(repair.equipment))
            self._detach_repair_from_batch(
                repair=repair,
                current_user=current_user,
            )

        if removed_labels:
            self._create_repair_message_record(
                repair=message_repair,
                author=current_user,
                payload=RepairMessageCreateRequest(
                    text="Из группы ремонта выведены приборы: "
                    + ", ".join(sorted(removed_labels))
                    + "."
                ),
                files=[],
            )
            self._record_event(
                category=EventCategory.REPAIR,
                action="repair_batch_items_removed",
                user=current_user,
                title=(
                    "Из группы ремонта "
                    f"«{anchor.batch_name or anchor.route_destination}» выведены приборы"
                ),
                description=f"Удалено приборов: {len(removed_labels)}.",
                equipment=message_repair.equipment,
                batch_key=anchor.batch_key,
            )

        self.session.commit()
        updated_repairs = self.repairs.list_active_by_batch_key(batch_key=normalized_batch_key)
        for repair in updated_repairs:
            self.session.refresh(repair)
        return updated_repairs

    def update_repair_milestones(
        self,
        *,
        equipment_id: int,
        payload: RepairMilestonesUpdateRequest,
        current_user: User,
    ) -> Repair:
        repair = self._get_active_repair(equipment_id=equipment_id)
        next_sent_to_repair_at = (
            payload.sent_to_repair_at
            if "sent_to_repair_at" in payload.model_fields_set
            else repair.sent_to_repair_at
        )
        if next_sent_to_repair_at is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Дата отправки в ремонт обязательна.",
            )
        _validate_repair_milestone_order(
            route_city=repair.route_city,
            route_destination=repair.route_destination,
            sent_to_repair_at=next_sent_to_repair_at,
            arrived_to_destination_at=(
                payload.arrived_to_destination_at
                if "arrived_to_destination_at" in payload.model_fields_set
                else repair.arrived_to_destination_at
            ),
            sent_from_repair_at=(
                payload.sent_from_repair_at
                if "sent_from_repair_at" in payload.model_fields_set
                else repair.sent_from_repair_at
            ),
            sent_from_irkutsk_at=(
                payload.sent_from_irkutsk_at
                if "sent_from_irkutsk_at" in payload.model_fields_set
                else repair.sent_from_irkutsk_at
            ),
            arrived_to_lensk_at=(
                payload.arrived_to_lensk_at
                if "arrived_to_lensk_at" in payload.model_fields_set
                else repair.arrived_to_lensk_at
            ),
            actually_received_at=(
                payload.actually_received_at
                if "actually_received_at" in payload.model_fields_set
                else repair.actually_received_at
            ),
            incoming_control_at=(
                payload.incoming_control_at
                if "incoming_control_at" in payload.model_fields_set
                else repair.incoming_control_at
            ),
            paid_at=(
                payload.paid_at
                if "paid_at" in payload.model_fields_set
                else repair.paid_at
            ),
        )
        if "sent_to_repair_at" in payload.model_fields_set:
            new_sent_to_repair_at = payload.sent_to_repair_at
            if new_sent_to_repair_at is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Дата отправки в ремонт обязательна.",
                )
            if repair.sent_to_repair_at != new_sent_to_repair_at:
                repair.sent_to_repair_at = new_sent_to_repair_at
                repair.repair_deadline_at = new_sent_to_repair_at + timedelta(days=100)
                self._create_repair_message_record(
                    repair=repair,
                    author=current_user,
                    payload=RepairMessageCreateRequest(
                        text=_build_repair_milestone_message(
                            current_user=current_user,
                            milestone_label=f"Отправлено в {repair.route_destination}",
                            milestone_date=new_sent_to_repair_at,
                        )
                    ),
                    files=[],
                )
        for field_name, milestone_label in _REPAIR_MILESTONE_LABELS:
            if field_name not in payload.model_fields_set:
                continue
            new_value = getattr(payload, field_name)
            current_value = getattr(repair, field_name)
            if current_value == new_value:
                continue
            setattr(repair, field_name, new_value)
            if new_value is not None:
                self._create_repair_message_record(
                    repair=repair,
                    author=current_user,
                    payload=RepairMessageCreateRequest(
                        text=_build_repair_milestone_message(
                            current_user=current_user,
                            milestone_label=milestone_label,
                            milestone_date=new_value,
                        )
                    ),
                    files=[],
                )

        changed_labels = _collect_changed_repair_milestone_labels(
            payload=payload,
            route_destination=repair.route_destination,
        )
        if changed_labels:
            self._record_event(
                category=EventCategory.REPAIR,
                action="repair_milestones_updated",
                user=current_user,
                title=f"Обновлены этапы ремонта «{repair.equipment.name}»",
                description="Изменено: " + ", ".join(changed_labels) + ".",
                equipment=repair.equipment,
                batch_key=repair.batch_key,
            )

        self.session.commit()
        self.session.refresh(repair)
        return repair

    def update_repair_batch_milestones(
        self,
        *,
        batch_key: str,
        payload: RepairMilestonesUpdateRequest,
        current_user: User,
    ) -> list[Repair]:
        normalized_batch_key = _normalize_required_text(
            batch_key,
            field_label="Repair batch key",
        )
        repairs = self.repairs.list_active_by_batch_key(batch_key=normalized_batch_key)
        if not repairs:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Активная группа ремонта не найдена.",
            )

        anchor = repairs[0]
        next_sent_to_repair_at = (
            payload.sent_to_repair_at
            if "sent_to_repair_at" in payload.model_fields_set
            else anchor.sent_to_repair_at
        )
        if next_sent_to_repair_at is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Дата отправки в ремонт обязательна.",
            )
        _validate_repair_milestone_order(
            route_city=anchor.route_city,
            route_destination=anchor.route_destination,
            sent_to_repair_at=next_sent_to_repair_at,
            arrived_to_destination_at=(
                payload.arrived_to_destination_at
                if "arrived_to_destination_at" in payload.model_fields_set
                else anchor.arrived_to_destination_at
            ),
            sent_from_repair_at=(
                payload.sent_from_repair_at
                if "sent_from_repair_at" in payload.model_fields_set
                else anchor.sent_from_repair_at
            ),
            sent_from_irkutsk_at=(
                payload.sent_from_irkutsk_at
                if "sent_from_irkutsk_at" in payload.model_fields_set
                else anchor.sent_from_irkutsk_at
            ),
            arrived_to_lensk_at=(
                payload.arrived_to_lensk_at
                if "arrived_to_lensk_at" in payload.model_fields_set
                else anchor.arrived_to_lensk_at
            ),
            actually_received_at=(
                payload.actually_received_at
                if "actually_received_at" in payload.model_fields_set
                else anchor.actually_received_at
            ),
            incoming_control_at=(
                payload.incoming_control_at
                if "incoming_control_at" in payload.model_fields_set
                else anchor.incoming_control_at
            ),
            paid_at=(
                payload.paid_at
                if "paid_at" in payload.model_fields_set
                else anchor.paid_at
            ),
        )

        if "sent_to_repair_at" in payload.model_fields_set:
            new_sent_to_repair_at = payload.sent_to_repair_at
            if new_sent_to_repair_at is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Дата отправки в ремонт обязательна.",
                )
            current_values = {item.sent_to_repair_at for item in repairs}
            if len(current_values) != 1 or new_sent_to_repair_at not in current_values:
                for repair in repairs:
                    repair.sent_to_repair_at = new_sent_to_repair_at
                    repair.repair_deadline_at = new_sent_to_repair_at + timedelta(days=100)
                self._create_repair_message_record(
                    repair=anchor,
                    author=current_user,
                    payload=RepairMessageCreateRequest(
                        text=_build_repair_milestone_message(
                            current_user=current_user,
                            milestone_label=f"Отправлено в {anchor.route_destination}",
                            milestone_date=new_sent_to_repair_at,
                        )
                    ),
                    files=[],
                )

        for field_name, milestone_label in _REPAIR_MILESTONE_LABELS:
            if field_name not in payload.model_fields_set:
                continue
            new_value = getattr(payload, field_name)
            current_values = {getattr(item, field_name) for item in repairs}
            if len(current_values) == 1 and new_value in current_values:
                continue
            for repair in repairs:
                setattr(repair, field_name, new_value)
            if new_value is not None:
                self._create_repair_message_record(
                    repair=anchor,
                    author=current_user,
                    payload=RepairMessageCreateRequest(
                        text=_build_repair_milestone_message(
                            current_user=current_user,
                            milestone_label=milestone_label,
                            milestone_date=new_value,
                        )
                    ),
                    files=[],
                )

        changed_labels = _collect_changed_repair_milestone_labels(
            payload=payload,
            route_destination=anchor.route_destination,
        )
        if changed_labels:
            self._record_event(
                category=EventCategory.REPAIR,
                action="repair_batch_milestones_updated",
                user=current_user,
                title=(
                    "Обновлены этапы группы ремонта "
                    f"«{anchor.batch_name or anchor.route_destination}»"
                ),
                description="Изменено: " + ", ".join(changed_labels) + ".",
                equipment=anchor.equipment,
                batch_key=anchor.batch_key,
            )

        self.session.commit()
        for repair in repairs:
            self.session.refresh(repair)
        return repairs

    def close_repair(
        self,
        *,
        equipment_id: int,
        current_user: User,
    ) -> Repair:
        repair = self._get_active_repair(equipment_id=equipment_id)
        if repair.paid_at is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Ремонт можно завершить только после даты оплаты.",
            )

        repair.closed_at = date.today()
        self._record_event(
            category=EventCategory.REPAIR,
            action="repair_closed",
            user=current_user,
            title=f"Ремонт «{repair.equipment.name}» завершен",
            description=_build_nonempty_description(
                [
                    _build_named_detail("Закрыт", self._format_sheet_date(repair.closed_at)),
                ]
            ),
            equipment=repair.equipment,
            batch_key=repair.batch_key,
        )
        self._sync_equipment_status(equipment=repair.equipment)
        self.session.commit()
        self.session.refresh(repair)
        return repair

    def close_repair_batch(
        self,
        *,
        batch_key: str,
        current_user: User,
    ) -> list[Repair]:
        normalized_batch_key = _normalize_required_text(
            batch_key,
            field_label="Repair batch key",
        )
        repairs = self.repairs.list_active_by_batch_key(batch_key=normalized_batch_key)
        if not repairs:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Активная группа ремонта не найдена.",
            )
        if any(repair.paid_at is None for repair in repairs):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Групповой ремонт можно завершить только после даты оплаты.",
            )

        closed_at = date.today()
        for repair in repairs:
            repair.closed_at = closed_at
        self._record_event(
            category=EventCategory.REPAIR,
            action="repair_batch_closed",
            user=current_user,
            title=(
                "Группа ремонта "
                f"«{repairs[0].batch_name or repairs[0].route_destination}» завершена"
            ),
            description=f"Закрыто приборов: {len(repairs)}.",
            equipment=repairs[0].equipment,
            batch_key=repairs[0].batch_key,
        )
        for repair in repairs:
            self._sync_equipment_status(equipment=repair.equipment)

        self.session.commit()
        for repair in repairs:
            self.session.refresh(repair)
        return repairs

    def create_verification(
        self,
        *,
        equipment_id: int,
        payload: VerificationCreateRequest,
        current_user: User,
        files: list[UploadedFilePayload] | None = None,
    ) -> Verification:
        equipment = self.get_equipment(equipment_id=equipment_id)
        if equipment.equipment_type != EquipmentType.SI:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Поверка доступна только для средств измерений.",
            )

        existing_verification = self.verifications.get_active_by_equipment_id(
            equipment_id=equipment.id
        )
        if existing_verification is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Для этого прибора уже есть активная поверка.",
            )

        sent_to_verification_at = payload.sent_to_verification_at
        verification = Verification(
            equipment_id=equipment.id,
            batch_key=_normalize_optional_text(payload.batch_key),
            batch_name=_normalize_optional_text(payload.batch_name),
            route_city=_normalize_required_text(
                payload.route_city,
                field_label="Verification route city",
            ),
            route_destination=_normalize_required_text(
                payload.route_destination,
                field_label="Verification route destination",
            ),
            sent_to_verification_at=sent_to_verification_at,
        )
        if self.repairs.get_active_by_equipment_id(equipment_id=equipment.id) is None:
            equipment.status = EquipmentStatus.IN_VERIFICATION
        self.verifications.add(verification)
        normalized_initial_text = _normalize_message_text(payload.initial_message_text)
        if normalized_initial_text is not None or files:
            self._create_verification_message_record(
                verification=verification,
                author=current_user,
                payload=VerificationMessageCreateRequest(text=normalized_initial_text),
                files=files or [],
            )
        self._record_event(
            category=EventCategory.VERIFICATION,
            action="verification_created",
            user=current_user,
            equipment=equipment,
            batch_key=verification.batch_key,
            title=f"Прибор «{equipment.name}» отправлен в поверку",
            description=_build_nonempty_description(
                [
                    _build_named_detail("Откуда", verification.route_city),
                    _build_named_detail("Куда", verification.route_destination),
                    _build_named_detail(
                        "Отправлено",
                        self._format_sheet_date(verification.sent_to_verification_at),
                    ),
                    _build_named_detail("Группа", verification.batch_name),
                ]
            ),
        )
        self.session.commit()
        self.session.refresh(verification)
        return verification

    def create_verification_batch(
        self,
        *,
        payload: VerificationBulkCreateRequest,
        current_user: User,
    ) -> list[Verification]:
        equipment_ids = list(dict.fromkeys(payload.equipment_ids))
        if not equipment_ids:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Нужно выбрать хотя бы один прибор для групповой поверки.",
            )

        batch_key = uuid4().hex
        batch_name = _normalize_required_text(
            payload.batch_name,
            field_label="Verification batch name",
        )
        created: list[Verification] = []
        for index, equipment_id in enumerate(equipment_ids):
            created.append(
                self.create_verification(
                    equipment_id=equipment_id,
                    payload=VerificationCreateRequest(
                        batch_key=batch_key,
                        batch_name=batch_name,
                        route_city=payload.route_city,
                        route_destination=payload.route_destination,
                        sent_to_verification_at=payload.sent_to_verification_at,
                        initial_message_text=payload.initial_message_text if index == 0 else None,
                    ),
                    current_user=current_user,
                    files=[],
                )
            )
        return created

    def update_verification_batch_items(
        self,
        *,
        batch_key: str,
        payload: ProcessBatchMembershipUpdateRequest,
        current_user: User,
    ) -> list[Verification]:
        normalized_batch_key = _normalize_required_text(
            batch_key,
            field_label="Verification batch key",
        )
        verifications = self.verifications.list_active_by_batch_key(batch_key=normalized_batch_key)
        if not verifications:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Активная группа поверки не найдена.",
            )

        anchor = verifications[0]
        existing_ids = {verification.equipment_id for verification in verifications}
        add_ids = [
            equipment_id
            for equipment_id in dict.fromkeys(payload.add_equipment_ids)
            if equipment_id not in existing_ids
        ]
        remove_ids = [
            equipment_id
            for equipment_id in dict.fromkeys(payload.remove_equipment_ids)
            if equipment_id in existing_ids
        ]

        if not add_ids and not remove_ids:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Нужно добавить или удалить хотя бы один прибор из группы поверки.",
            )
        if len(remove_ids) >= len(verifications):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Нельзя удалить из группы поверки все приборы сразу.",
            )

        added_labels: list[str] = []
        removed_labels: list[str] = []

        for equipment_id in add_ids:
            new_verification = self._create_verification_from_batch_anchor(
                anchor=anchor,
                equipment_id=equipment_id,
            )
            added_labels.append(_build_equipment_batch_member_label(new_verification.equipment))

        if added_labels:
            self._create_verification_message_record(
                verification=anchor,
                author=current_user,
                payload=VerificationMessageCreateRequest(
                    text="В группу поверки добавлены приборы: "
                    + ", ".join(sorted(added_labels))
                    + "."
                ),
                files=[],
            )
            self._record_event(
                category=EventCategory.VERIFICATION,
                action="verification_batch_items_added",
                user=current_user,
                title=(
                    "В группу поверки "
                    f"«{anchor.batch_name or anchor.route_destination}» добавлены приборы"
                ),
                description=f"Добавлено СИ: {len(added_labels)}.",
                equipment=anchor.equipment,
                batch_key=anchor.batch_key,
            )

        message_verification = next(
            (item for item in verifications if item.equipment_id not in remove_ids),
            anchor,
        )
        for verification in verifications:
            if verification.equipment_id not in remove_ids:
                continue
            removed_labels.append(_build_equipment_batch_member_label(verification.equipment))
            self._detach_verification_from_batch(
                verification=verification,
                current_user=current_user,
            )

        if removed_labels:
            self._create_verification_message_record(
                verification=message_verification,
                author=current_user,
                payload=VerificationMessageCreateRequest(
                    text="Из группы поверки выведены приборы: "
                    + ", ".join(sorted(removed_labels))
                    + "."
                ),
                files=[],
            )
            self._record_event(
                category=EventCategory.VERIFICATION,
                action="verification_batch_items_removed",
                user=current_user,
                title=(
                    "Из группы поверки "
                    f"«{anchor.batch_name or anchor.route_destination}» выведены приборы"
                ),
                description=f"Удалено СИ: {len(removed_labels)}.",
                equipment=message_verification.equipment,
                batch_key=anchor.batch_key,
            )

        self.session.commit()
        updated_verifications = self.verifications.list_active_by_batch_key(
            batch_key=normalized_batch_key
        )
        for verification in updated_verifications:
            self.session.refresh(verification)
        return updated_verifications

    def close_verification(
        self,
        *,
        equipment_id: int,
        current_user: User,
    ) -> Verification:
        verification = self._get_active_verification(equipment_id=equipment_id)
        verification.closed_at = date.today()
        self._record_event(
            category=EventCategory.VERIFICATION,
            action="verification_closed",
            user=current_user,
            title=f"Поверка «{verification.equipment.name}» завершена",
            description=_build_nonempty_description(
                [
                    _build_named_detail("Закрыта", self._format_sheet_date(verification.closed_at)),
                ]
            ),
            equipment=verification.equipment,
            batch_key=verification.batch_key,
        )
        self._sync_equipment_status(equipment=verification.equipment)
        self.session.commit()
        self.session.refresh(verification)
        return verification

    def close_verification_batch(
        self,
        *,
        batch_key: str,
        current_user: User,
    ) -> list[Verification]:
        normalized_batch_key = _normalize_required_text(
            batch_key,
            field_label="Verification batch key",
        )
        verifications = self.verifications.list_active_by_batch_key(batch_key=normalized_batch_key)
        if not verifications:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Активная группа поверки не найдена.",
            )

        closed_at = date.today()
        for verification in verifications:
            verification.closed_at = closed_at
        self._record_event(
            category=EventCategory.VERIFICATION,
            action="verification_batch_closed",
            user=current_user,
            title=(
                "Группа поверки "
                f"«{verifications[0].batch_name or verifications[0].route_destination}» завершена"
            ),
            description=f"Закрыто СИ: {len(verifications)}.",
            equipment=verifications[0].equipment,
            batch_key=verifications[0].batch_key,
        )
        for verification in verifications:
            self._sync_equipment_status(equipment=verification.equipment)

        self.session.commit()
        for verification in verifications:
            self.session.refresh(verification)
        return verifications

    def update_verification_milestones(
        self,
        *,
        equipment_id: int,
        payload: VerificationMilestonesUpdateRequest,
        current_user: User,
    ) -> Verification:
        verification = self._get_active_verification(equipment_id=equipment_id)
        _validate_verification_milestone_order(
            route_destination=verification.route_destination,
            sent_to_verification_at=verification.sent_to_verification_at,
            received_at_destination_at=(
                payload.received_at_destination_at
                if "received_at_destination_at" in payload.model_fields_set
                else verification.received_at_destination_at
            ),
            handed_to_csm_at=(
                payload.handed_to_csm_at
                if "handed_to_csm_at" in payload.model_fields_set
                else verification.handed_to_csm_at
            ),
            verification_completed_at=(
                payload.verification_completed_at
                if "verification_completed_at" in payload.model_fields_set
                else verification.verification_completed_at
            ),
            picked_up_from_csm_at=(
                payload.picked_up_from_csm_at
                if "picked_up_from_csm_at" in payload.model_fields_set
                else verification.picked_up_from_csm_at
            ),
            shipped_back_at=(
                payload.shipped_back_at
                if "shipped_back_at" in payload.model_fields_set
                else verification.shipped_back_at
            ),
            returned_from_verification_at=(
                payload.returned_from_verification_at
                if "returned_from_verification_at" in payload.model_fields_set
                else verification.returned_from_verification_at
            ),
        )
        for field_name, milestone_label in _VERIFICATION_MILESTONE_LABELS:
            if field_name not in payload.model_fields_set:
                continue
            new_value = getattr(payload, field_name)
            current_value = getattr(verification, field_name)
            if current_value == new_value:
                continue
            setattr(verification, field_name, new_value)
            if new_value is not None:
                self._create_verification_message_record(
                    verification=verification,
                    author=current_user,
                    payload=VerificationMessageCreateRequest(
                        text=_build_verification_milestone_message(
                            current_user=current_user,
                            milestone_label=milestone_label,
                            milestone_date=new_value,
                            route_destination=verification.route_destination,
                        )
                    ),
                    files=[],
                )

        changed_labels = _collect_changed_verification_milestone_labels(
            payload=payload,
            route_destination=verification.route_destination,
        )
        if changed_labels:
            self._record_event(
                category=EventCategory.VERIFICATION,
                action="verification_milestones_updated",
                user=current_user,
                title=f"Обновлены этапы поверки «{verification.equipment.name}»",
                description="Изменено: " + ", ".join(changed_labels) + ".",
                equipment=verification.equipment,
                batch_key=verification.batch_key,
            )

        self.session.commit()
        self.session.refresh(verification)
        return verification

    def update_verification_batch_milestones(
        self,
        *,
        batch_key: str,
        payload: VerificationMilestonesUpdateRequest,
        current_user: User,
    ) -> list[Verification]:
        normalized_batch_key = _normalize_required_text(
            batch_key,
            field_label="Verification batch key",
        )
        verifications = self.verifications.list_active_by_batch_key(batch_key=normalized_batch_key)
        if not verifications:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Активная группа поверки не найдена.",
            )

        anchor = verifications[0]
        _validate_verification_milestone_order(
            route_destination=anchor.route_destination,
            sent_to_verification_at=anchor.sent_to_verification_at,
            received_at_destination_at=(
                payload.received_at_destination_at
                if "received_at_destination_at" in payload.model_fields_set
                else anchor.received_at_destination_at
            ),
            handed_to_csm_at=(
                payload.handed_to_csm_at
                if "handed_to_csm_at" in payload.model_fields_set
                else anchor.handed_to_csm_at
            ),
            verification_completed_at=(
                payload.verification_completed_at
                if "verification_completed_at" in payload.model_fields_set
                else anchor.verification_completed_at
            ),
            picked_up_from_csm_at=(
                payload.picked_up_from_csm_at
                if "picked_up_from_csm_at" in payload.model_fields_set
                else anchor.picked_up_from_csm_at
            ),
            shipped_back_at=(
                payload.shipped_back_at
                if "shipped_back_at" in payload.model_fields_set
                else anchor.shipped_back_at
            ),
            returned_from_verification_at=(
                payload.returned_from_verification_at
                if "returned_from_verification_at" in payload.model_fields_set
                else anchor.returned_from_verification_at
            ),
        )
        for field_name, milestone_label in _VERIFICATION_MILESTONE_LABELS:
            if field_name not in payload.model_fields_set:
                continue
            new_value = getattr(payload, field_name)
            current_values = {getattr(item, field_name) for item in verifications}
            if len(current_values) == 1 and new_value in current_values:
                continue
            for verification in verifications:
                setattr(verification, field_name, new_value)
            if new_value is not None:
                self._create_verification_message_record(
                    verification=anchor,
                    author=current_user,
                    payload=VerificationMessageCreateRequest(
                        text=_build_verification_milestone_message(
                            current_user=current_user,
                            milestone_label=milestone_label,
                            milestone_date=new_value,
                            route_destination=anchor.route_destination,
                        )
                    ),
                    files=[],
                )

        changed_labels = _collect_changed_verification_milestone_labels(
            payload=payload,
            route_destination=anchor.route_destination,
        )
        if changed_labels:
            self._record_event(
                category=EventCategory.VERIFICATION,
                action="verification_batch_milestones_updated",
                user=current_user,
                title=(
                    "Обновлены этапы группы поверки "
                    f"«{anchor.batch_name or anchor.route_destination}»"
                ),
                description="Изменено: " + ", ".join(changed_labels) + ".",
                equipment=anchor.equipment,
                batch_key=anchor.batch_key,
            )

        self.session.commit()
        for verification in verifications:
            self.session.refresh(verification)
        return verifications

    def list_active_repair_messages(self, *, equipment_id: int) -> list[RepairMessage]:
        repair = self._get_active_repair(equipment_id=equipment_id)
        if repair.batch_key:
            return self.repair_messages.list_by_batch_key(batch_key=repair.batch_key)
        return self.repair_messages.list_by_repair(repair_id=repair.id)

    def list_active_verification_messages(
        self,
        *,
        equipment_id: int,
    ) -> list[VerificationMessage]:
        verification = self._get_active_verification(equipment_id=equipment_id)
        if verification.batch_key:
            return self.verification_messages.list_by_batch_key(
                batch_key=verification.batch_key
            )
        return self.verification_messages.list_by_verification(
            verification_id=verification.id
        )

    def create_repair_message(
        self,
        *,
        equipment_id: int,
        payload: RepairMessageCreateRequest,
        author: User,
        files: list[UploadedFilePayload] | None = None,
    ) -> RepairMessage:
        repair = self._get_active_repair(equipment_id=equipment_id)
        message = self._create_repair_message_record(
            repair=repair,
            author=author,
            payload=payload,
            files=files or [],
        )
        self._record_event(
            category=EventCategory.REPAIR,
            action="repair_message_created",
            user=author,
            title=f"Добавлено сообщение по ремонту «{repair.equipment.name}»",
            description=_build_message_event_description(
                text=payload.text,
                files=files or [],
            ),
            equipment=repair.equipment,
            batch_key=repair.batch_key,
        )
        self.session.commit()
        self.session.refresh(message)
        return message

    def delete_repair_message(
        self,
        *,
        equipment_id: int,
        message_id: int,
        current_user: User,
    ) -> None:
        repair = self._get_active_repair(equipment_id=equipment_id)
        message = self._get_repair_message(
            repair=repair,
            message_id=message_id,
        )
        self._assert_repair_message_owner(
            message=message,
            current_user=current_user,
        )

        attachment_paths = [
            settings.attachment_storage_path / attachment.storage_path
            for attachment in message.attachments
        ]
        message_preview = message.text
        for attachment in message.attachments:
            self.repair_message_attachments.delete(attachment)
        self.repair_messages.delete(message)
        self._record_event(
            category=EventCategory.REPAIR,
            action="repair_message_deleted",
            user=current_user,
            title=f"Удалено сообщение по ремонту «{repair.equipment.name}»",
            description=_build_preview_description(message_preview),
            equipment=repair.equipment,
            batch_key=repair.batch_key,
        )
        self.session.commit()

        for file_path in attachment_paths:
            if file_path.exists() and file_path.is_file():
                file_path.unlink()

    def create_verification_message(
        self,
        *,
        equipment_id: int,
        payload: VerificationMessageCreateRequest,
        author: User,
        files: list[UploadedFilePayload] | None = None,
    ) -> VerificationMessage:
        verification = self._get_active_verification(equipment_id=equipment_id)
        message = self._create_verification_message_record(
            verification=verification,
            author=author,
            payload=payload,
            files=files or [],
        )
        self._record_event(
            category=EventCategory.VERIFICATION,
            action="verification_message_created",
            user=author,
            title=f"Добавлено сообщение по поверке «{verification.equipment.name}»",
            description=_build_message_event_description(
                text=payload.text,
                files=files or [],
            ),
            equipment=verification.equipment,
            batch_key=verification.batch_key,
        )
        self.session.commit()
        self.session.refresh(message)
        return message

    def delete_verification_message(
        self,
        *,
        equipment_id: int,
        message_id: int,
        current_user: User,
    ) -> None:
        verification = self._get_active_verification(equipment_id=equipment_id)
        message = self._get_verification_message(
            verification_id=verification.id,
            message_id=message_id,
        )
        self._assert_verification_message_owner(
            message=message,
            current_user=current_user,
        )

        attachment_paths = [
            settings.attachment_storage_path / attachment.storage_path
            for attachment in message.attachments
        ]
        message_preview = message.text
        for attachment in message.attachments:
            self.verification_message_attachments.delete(attachment)
        self.verification_messages.delete(message)
        self._record_event(
            category=EventCategory.VERIFICATION,
            action="verification_message_deleted",
            user=current_user,
            title=f"Удалено сообщение по поверке «{verification.equipment.name}»",
            description=_build_preview_description(message_preview),
            equipment=verification.equipment,
            batch_key=verification.batch_key,
        )
        self.session.commit()

        for file_path in attachment_paths:
            if file_path.exists() and file_path.is_file():
                file_path.unlink()

    def export_verification_archive_zip(
        self,
        *,
        verification_id: int,
    ) -> tuple[str, bytes]:
        verification = self.verifications.get_by_id(verification_id)
        if verification is None or verification.closed_at is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Архив поверки не найден.",
            )

        if verification.batch_key:
            messages = self.verification_messages.list_by_batch_key(
                batch_key=verification.batch_key
            )
            archive_name = _build_verification_archive_name(
                label=verification.batch_name or "Групповая поверка",
                start_date=verification.sent_to_verification_at,
                end_date=verification.closed_at,
            )
        else:
            messages = self.verification_messages.list_by_verification(
                verification_id=verification.id
            )
            equipment = verification.equipment
            folder_name = "Без папки"
            if equipment is not None and equipment.folder_id is not None:
                folder = self.folders.get_by_id(equipment.folder_id)
                if folder is not None:
                    folder_name = folder.name
            archive_name = _build_verification_archive_name(
                label=folder_name,
                start_date=verification.sent_to_verification_at,
                end_date=verification.closed_at,
            )

        buffer = BytesIO()
        with ZipFile(buffer, mode="w", compression=ZIP_DEFLATED) as zip_file:
            transcript_lines: list[str] = []
            for index, message in enumerate(messages, start=1):
                created_at = message.created_at.strftime("%d.%m.%Y %H:%M")
                transcript_lines.append(f"[{created_at}] {message.author_display_name}")
                if message.text:
                    transcript_lines.append(message.text)
                if message.attachments:
                    transcript_lines.append(
                        "Вложения: "
                        + ", ".join(attachment.file_name for attachment in message.attachments)
                    )
                transcript_lines.append("")

                for attachment_index, attachment in enumerate(
                    message.attachments,
                    start=1,
                ):
                    file_path = settings.attachment_storage_path / attachment.storage_path
                    if file_path.exists() and file_path.is_file():
                        zip_file.write(
                            file_path,
                            arcname=(
                                f"files/{index:03d}_{attachment_index:02d}_{attachment.file_name}"
                            ),
                        )

            transcript_content = "\n".join(transcript_lines).strip()
            if not transcript_content:
                transcript_content = "Диалог поверки пуст."

            zip_file.writestr(
                "dialog.txt",
                transcript_content,
            )

        return f"{archive_name}.zip", buffer.getvalue()

    def export_repair_archive_zip(
        self,
        *,
        repair_id: int,
    ) -> tuple[str, bytes]:
        repair = self.repairs.get_by_id(repair_id)
        if repair is None or repair.closed_at is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Архив ремонта не найден.",
            )

        if repair.batch_key:
            messages = self.repair_messages.list_by_batch_key(batch_key=repair.batch_key)
            archive_name = _build_repair_archive_name(
                folder_name=repair.batch_name or "Групповой ремонт",
                start_date=repair.sent_to_repair_at,
                end_date=repair.closed_at,
            )
        else:
            messages = self.repair_messages.list_by_repair(repair_id=repair.id)
            equipment = repair.equipment
            folder_name = "Без папки"
            if equipment is not None and equipment.folder_id is not None:
                folder = self.folders.get_by_id(equipment.folder_id)
                if folder is not None:
                    folder_name = folder.name
            archive_name = _build_repair_archive_name(
                folder_name=folder_name,
                start_date=repair.sent_to_repair_at,
                end_date=repair.closed_at,
            )

        buffer = BytesIO()
        with ZipFile(buffer, mode="w", compression=ZIP_DEFLATED) as zip_file:
            transcript_lines: list[str] = []
            for index, message in enumerate(messages, start=1):
                created_at = message.created_at.strftime("%d.%m.%Y %H:%M")
                transcript_lines.append(f"[{created_at}] {message.author_display_name}")
                if message.text:
                    transcript_lines.append(message.text)
                if message.attachments:
                    transcript_lines.append(
                        "Вложения: "
                        + ", ".join(attachment.file_name for attachment in message.attachments)
                    )
                transcript_lines.append("")

                for attachment_index, attachment in enumerate(
                    message.attachments,
                    start=1,
                ):
                    file_path = settings.attachment_storage_path / attachment.storage_path
                    if file_path.exists() and file_path.is_file():
                        zip_file.write(
                            file_path,
                            arcname=(
                                f"files/{index:03d}_{attachment_index:02d}_{attachment.file_name}"
                            ),
                        )

            transcript_content = "\n".join(transcript_lines).strip()
            if not transcript_content:
                transcript_content = "Диалог ремонта пуст."

            zip_file.writestr(
                "dialog.txt",
                transcript_content,
            )

        return f"{archive_name}.zip", buffer.getvalue()

    def get_repair_message_attachment_file(
        self,
        *,
        equipment_id: int,
        message_id: int,
        attachment_id: int,
    ) -> tuple[RepairMessageAttachment, Path]:
        repair = self._get_active_repair(equipment_id=equipment_id)
        message = self._get_repair_message(repair=repair, message_id=message_id)
        attachment = self._get_repair_message_attachment(
            message_id=message.id,
            attachment_id=attachment_id,
        )
        file_path = settings.attachment_storage_path / attachment.storage_path
        if not file_path.exists() or not file_path.is_file():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Repair attachment file is missing.",
            )
        return attachment, file_path

    def get_verification_message_attachment_file(
        self,
        *,
        equipment_id: int,
        message_id: int,
        attachment_id: int,
    ) -> tuple[VerificationMessageAttachment, Path]:
        verification = self._get_active_verification(equipment_id=equipment_id)
        message = self._get_verification_message(
            verification_id=verification.id,
            message_id=message_id,
        )
        attachment = self._get_verification_message_attachment(
            message_id=message.id,
            attachment_id=attachment_id,
        )
        file_path = settings.attachment_storage_path / attachment.storage_path
        if not file_path.exists() or not file_path.is_file():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Verification attachment file is missing.",
            )
        return attachment, file_path

    def delete_equipment(self, *, equipment_id: int, current_user: User | None = None) -> None:
        equipment = self.get_equipment(equipment_id=equipment_id)
        if current_user is not None:
            self._record_equipment_event(
                action="equipment_deleted",
                user=current_user,
                equipment=equipment,
                title=f"Удален прибор «{equipment.name}»",
            )
        self.equipment.delete(equipment)
        self.session.commit()

    def delete_equipment_batch(
        self,
        *,
        equipment_ids: list[int],
        current_user: User | None = None,
    ) -> None:
        normalized_ids = list(dict.fromkeys(equipment_ids))
        if not normalized_ids:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Нужно выбрать хотя бы один прибор для удаления.",
            )

        for equipment_id in normalized_ids:
            equipment = self.get_equipment(equipment_id=equipment_id)
            if current_user is not None:
                self._record_equipment_event(
                    action="equipment_deleted",
                    user=current_user,
                    equipment=equipment,
                    title=f"Удален прибор «{equipment.name}»",
                )
            self.equipment.delete(equipment)

        self.session.commit()

    def list_attachments(self, *, equipment_id: int) -> list[EquipmentAttachment]:
        self.get_equipment(equipment_id=equipment_id)
        return self.attachments.list_by_equipment(equipment_id=equipment_id)

    def create_attachment(
        self,
        *,
        equipment_id: int,
        uploader: User,
        file_name: str | None,
        content_type: str | None,
        content: bytes,
    ) -> EquipmentAttachment:
        equipment = self.get_equipment(equipment_id=equipment_id)
        normalized_file_name = _normalize_attachment_file_name(file_name)
        if not content:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Attachment file must not be empty.",
            )

        storage_dir = settings.attachment_storage_path / str(equipment.id)
        storage_dir.mkdir(parents=True, exist_ok=True)
        storage_name = f"{uuid4().hex}{Path(normalized_file_name).suffix.lower()}"
        file_path = storage_dir / storage_name
        file_path.write_bytes(content)

        relative_storage_path = str(file_path.relative_to(settings.attachment_storage_path))
        attachment = EquipmentAttachment(
            equipment_id=equipment.id,
            uploaded_by_user_id=uploader.id,
            uploaded_by_display_name=_format_user_display_name(uploader),
            file_name=normalized_file_name,
            file_mime_type=_normalize_optional_text(content_type),
            file_size=len(content),
            storage_path=relative_storage_path,
        )
        self.attachments.add(attachment)
        self._record_equipment_event(
            action="attachment_created",
            user=uploader,
            equipment=equipment,
            title=f"Добавлено вложение к прибору «{equipment.name}»",
            description=normalized_file_name,
        )
        self.session.commit()
        self.session.refresh(attachment)
        return attachment

    def get_attachment_file(
        self,
        *,
        equipment_id: int,
        attachment_id: int,
    ) -> tuple[EquipmentAttachment, Path]:
        attachment = self._get_attachment(equipment_id=equipment_id, attachment_id=attachment_id)
        file_path = settings.attachment_storage_path / attachment.storage_path
        if not file_path.exists() or not file_path.is_file():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Attachment file is missing.",
            )
        return attachment, file_path

    def delete_attachment(
        self,
        *,
        equipment_id: int,
        attachment_id: int,
        current_user: User,
    ) -> None:
        attachment = self._get_attachment(equipment_id=equipment_id, attachment_id=attachment_id)
        if current_user.role not in {UserRole.ADMINISTRATOR, UserRole.MKAIR}:
            if attachment.uploaded_by_user_id != current_user.id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="You cannot delete this attachment.",
                )

        file_path = settings.attachment_storage_path / attachment.storage_path
        attachment_name = attachment.file_name
        equipment = self.get_equipment(equipment_id=equipment_id)
        self.attachments.delete(attachment)
        self._record_equipment_event(
            action="attachment_deleted",
            user=current_user,
            equipment=equipment,
            title=f"Удалено вложение у прибора «{equipment.name}»",
            description=attachment_name,
        )
        self.session.commit()
        if file_path.exists() and file_path.is_file():
            file_path.unlink()

    def list_comments(self, *, equipment_id: int) -> list[EquipmentComment]:
        self.get_equipment(equipment_id=equipment_id)
        return self.comments.list_by_equipment(equipment_id=equipment_id)

    def create_comment(
        self,
        *,
        equipment_id: int,
        payload: EquipmentCommentCreateRequest,
        author: User,
    ) -> EquipmentComment:
        equipment = self.get_equipment(equipment_id=equipment_id)
        normalized_text = _normalize_comment_text(payload.text)
        comment = EquipmentComment(
            equipment_id=equipment.id,
            author_user_id=author.id,
            author_display_name=_format_user_display_name(author),
            text=normalized_text,
        )
        self.comments.add(comment)
        self._record_equipment_event(
            action="comment_created",
            user=author,
            equipment=equipment,
            title=f"Добавлен комментарий к прибору «{equipment.name}»",
            description=_build_preview_description(comment.text),
        )
        self.session.commit()
        self.session.refresh(comment)
        return comment

    def update_comment(
        self,
        *,
        equipment_id: int,
        comment_id: int,
        payload: EquipmentCommentUpdateRequest,
        current_user: User,
    ) -> EquipmentComment:
        comment = self._get_comment(equipment_id=equipment_id, comment_id=comment_id)
        self._assert_comment_owner(comment=comment, current_user=current_user)
        comment.text = _normalize_comment_text(payload.text)
        equipment = self.get_equipment(equipment_id=equipment_id)
        self._record_equipment_event(
            action="comment_updated",
            user=current_user,
            equipment=equipment,
            title=f"Обновлен комментарий к прибору «{equipment.name}»",
            description=_build_preview_description(comment.text),
        )
        self.session.commit()
        self.session.refresh(comment)
        return comment

    def delete_comment(
        self,
        *,
        equipment_id: int,
        comment_id: int,
        current_user: User,
    ) -> None:
        comment = self._get_comment(equipment_id=equipment_id, comment_id=comment_id)
        self._assert_comment_owner(comment=comment, current_user=current_user)
        equipment = self.get_equipment(equipment_id=equipment_id)
        self._record_equipment_event(
            action="comment_deleted",
            user=current_user,
            equipment=equipment,
            title=f"Удален комментарий у прибора «{equipment.name}»",
            description=_build_preview_description(comment.text),
        )
        self.comments.delete(comment)
        self.session.commit()

    def _record_equipment_event(
        self,
        *,
        action: str,
        user: User,
        equipment: Equipment,
        title: str,
        description: str | None = None,
        batch_key: str | None = None,
    ) -> None:
        self._record_event(
            category=EventCategory.EQUIPMENT,
            action=action,
            user=user,
            title=title,
            description=description,
            equipment=equipment,
            batch_key=batch_key,
        )

    def _record_event(
        self,
        *,
        category: EventCategory,
        action: str,
        user: User | None,
        title: str,
        description: str | None = None,
        equipment: Equipment | None = None,
        equipment_id: int | None = None,
        equipment_name: str | None = None,
        equipment_modification: str | None = None,
        equipment_serial_number: str | None = None,
        folder_id: int | None = None,
        folder_name: str | None = None,
        batch_key: str | None = None,
    ) -> None:
        resolved_equipment_id = equipment.id if equipment is not None else equipment_id
        resolved_equipment_name = equipment.name if equipment is not None else equipment_name
        resolved_equipment_modification = (
            equipment.modification if equipment is not None else equipment_modification
        )
        resolved_equipment_serial_number = (
            equipment.serial_number if equipment is not None else equipment_serial_number
        )
        resolved_folder_id = folder_id
        resolved_folder_name = folder_name

        if equipment is not None:
            resolved_folder_id = equipment.folder_id
            if resolved_folder_name is None and equipment.folder_id is not None:
                folder = self.folders.get_by_id(equipment.folder_id)
                if folder is not None:
                    resolved_folder_name = folder.name

        event = EventLog(
            category=category,
            action=action,
            title=title,
            description=description,
            user_id=user.id if user is not None else None,
            user_display_name=(
                _format_user_display_name(user) if user is not None else "Система"
            ),
            equipment_id=resolved_equipment_id,
            equipment_name=resolved_equipment_name,
            equipment_modification=resolved_equipment_modification,
            equipment_serial_number=resolved_equipment_serial_number,
            folder_id=resolved_folder_id,
            folder_name=resolved_folder_name,
            batch_key=_normalize_optional_text(batch_key),
        )
        self.events.add(event)

    def _create_repair_from_batch_anchor(
        self,
        *,
        anchor: Repair,
        equipment_id: int,
    ) -> Repair:
        equipment = self.get_equipment(equipment_id=equipment_id)
        existing_repair = self.repairs.get_active_by_equipment_id(equipment_id=equipment.id)
        if existing_repair is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Для прибора «{equipment.name}» уже есть активный ремонт.",
            )

        repair = Repair(
            equipment_id=equipment.id,
            batch_key=anchor.batch_key,
            batch_name=anchor.batch_name,
            route_city=anchor.route_city,
            route_destination=anchor.route_destination,
            sent_to_repair_at=anchor.sent_to_repair_at,
            repair_deadline_at=anchor.repair_deadline_at,
            arrived_to_destination_at=anchor.arrived_to_destination_at,
            sent_from_repair_at=anchor.sent_from_repair_at,
            sent_from_irkutsk_at=anchor.sent_from_irkutsk_at,
            arrived_to_lensk_at=anchor.arrived_to_lensk_at,
            actually_received_at=anchor.actually_received_at,
            incoming_control_at=anchor.incoming_control_at,
            paid_at=anchor.paid_at,
        )
        equipment.status = EquipmentStatus.IN_REPAIR
        self.repairs.add(repair)
        self.session.flush()
        self.session.refresh(repair)
        return repair

    def _create_verification_from_batch_anchor(
        self,
        *,
        anchor: Verification,
        equipment_id: int,
    ) -> Verification:
        equipment = self.get_equipment(equipment_id=equipment_id)
        if equipment.equipment_type != EquipmentType.SI:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Прибор «{equipment.name}» нельзя добавить в поверку: это не СИ.",
            )
        existing_verification = self.verifications.get_active_by_equipment_id(
            equipment_id=equipment.id
        )
        if existing_verification is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Для прибора «{equipment.name}» уже есть активная поверка.",
            )

        verification = Verification(
            equipment_id=equipment.id,
            batch_key=anchor.batch_key,
            batch_name=anchor.batch_name,
            route_city=anchor.route_city,
            route_destination=anchor.route_destination,
            sent_to_verification_at=anchor.sent_to_verification_at,
            received_at_destination_at=anchor.received_at_destination_at,
            handed_to_csm_at=anchor.handed_to_csm_at,
            verification_completed_at=anchor.verification_completed_at,
            picked_up_from_csm_at=anchor.picked_up_from_csm_at,
            shipped_back_at=anchor.shipped_back_at,
            returned_from_verification_at=anchor.returned_from_verification_at,
        )
        if self.repairs.get_active_by_equipment_id(equipment_id=equipment.id) is None:
            equipment.status = EquipmentStatus.IN_VERIFICATION
        self.verifications.add(verification)
        self.session.flush()
        self.session.refresh(verification)
        return verification

    def _detach_repair_from_batch(
        self,
        *,
        repair: Repair,
        current_user: User,
    ) -> None:
        batch_key = repair.batch_key
        batch_name = repair.batch_name or "Групповой ремонт"
        if batch_key is None:
            return

        self._clone_repair_batch_history_to_repair(
            repair=repair,
            batch_key=batch_key,
        )
        member_label = _build_equipment_batch_member_label(repair.equipment)
        repair.batch_key = None
        repair.batch_name = None
        self._create_repair_message_record(
            repair=repair,
            author=current_user,
            payload=RepairMessageCreateRequest(
                text=f"Прибор выведен из группы «{batch_name}»: {member_label}."
            ),
            files=[],
        )

    def _detach_verification_from_batch(
        self,
        *,
        verification: Verification,
        current_user: User,
    ) -> None:
        batch_key = verification.batch_key
        batch_name = verification.batch_name or "Групповая поверка"
        if batch_key is None:
            return

        self._clone_verification_batch_history_to_verification(
            verification=verification,
            batch_key=batch_key,
        )
        member_label = _build_equipment_batch_member_label(verification.equipment)
        verification.batch_key = None
        verification.batch_name = None
        self._create_verification_message_record(
            verification=verification,
            author=current_user,
            payload=VerificationMessageCreateRequest(
                text=f"Прибор выведен из группы «{batch_name}»: {member_label}."
            ),
            files=[],
        )

    def _clone_repair_batch_history_to_repair(
        self,
        *,
        repair: Repair,
        batch_key: str,
    ) -> None:
        messages = self.repair_messages.list_by_batch_key(batch_key=batch_key)
        for message in messages:
            if message.repair_id == repair.id:
                continue
            cloned_message = RepairMessage(
                repair_id=repair.id,
                batch_key=None,
                author_user_id=message.author_user_id,
                author_display_name=message.author_display_name,
                text=message.text,
                created_at=message.created_at,
            )
            self.repair_messages.add(cloned_message)
            for attachment in message.attachments:
                self._copy_repair_message_attachment(
                    repair=repair,
                    message=cloned_message,
                    source=attachment,
                )

    def _clone_verification_batch_history_to_verification(
        self,
        *,
        verification: Verification,
        batch_key: str,
    ) -> None:
        messages = self.verification_messages.list_by_batch_key(batch_key=batch_key)
        for message in messages:
            if message.verification_id == verification.id:
                continue
            cloned_message = VerificationMessage(
                verification_id=verification.id,
                batch_key=None,
                author_user_id=message.author_user_id,
                author_display_name=message.author_display_name,
                text=message.text,
                created_at=message.created_at,
            )
            self.verification_messages.add(cloned_message)
            for attachment in message.attachments:
                self._copy_verification_message_attachment(
                    verification=verification,
                    message=cloned_message,
                    source=attachment,
                )

    def _copy_repair_message_attachment(
        self,
        *,
        repair: Repair,
        message: RepairMessage,
        source: RepairMessageAttachment,
    ) -> RepairMessageAttachment | None:
        source_path = settings.attachment_storage_path / source.storage_path
        if not source_path.exists() or not source_path.is_file():
            return None

        storage_dir = (
            settings.attachment_storage_path
            / "repair-messages"
            / str(repair.id)
            / str(message.id)
        )
        storage_dir.mkdir(parents=True, exist_ok=True)
        storage_name = f"{uuid4().hex}{Path(source.file_name).suffix.lower()}"
        file_path = storage_dir / storage_name
        file_path.write_bytes(source_path.read_bytes())

        attachment = RepairMessageAttachment(
            repair_message_id=message.id,
            uploaded_by_user_id=source.uploaded_by_user_id,
            uploaded_by_display_name=source.uploaded_by_display_name,
            file_name=source.file_name,
            file_mime_type=source.file_mime_type,
            file_size=source.file_size,
            storage_path=str(file_path.relative_to(settings.attachment_storage_path)),
            created_at=source.created_at,
        )
        self.repair_message_attachments.add(attachment)
        return attachment

    def _copy_verification_message_attachment(
        self,
        *,
        verification: Verification,
        message: VerificationMessage,
        source: VerificationMessageAttachment,
    ) -> VerificationMessageAttachment | None:
        source_path = settings.attachment_storage_path / source.storage_path
        if not source_path.exists() or not source_path.is_file():
            return None

        storage_dir = (
            settings.attachment_storage_path
            / "verification-messages"
            / str(verification.id)
            / str(message.id)
        )
        storage_dir.mkdir(parents=True, exist_ok=True)
        storage_name = f"{uuid4().hex}{Path(source.file_name).suffix.lower()}"
        file_path = storage_dir / storage_name
        file_path.write_bytes(source_path.read_bytes())

        attachment = VerificationMessageAttachment(
            verification_message_id=message.id,
            uploaded_by_user_id=source.uploaded_by_user_id,
            uploaded_by_display_name=source.uploaded_by_display_name,
            file_name=source.file_name,
            file_mime_type=source.file_mime_type,
            file_size=source.file_size,
            storage_path=str(file_path.relative_to(settings.attachment_storage_path)),
            created_at=source.created_at,
        )
        self.verification_message_attachments.add(attachment)
        return attachment

    def _create_repair_message_record(
        self,
        *,
        repair: Repair,
        author: User,
        payload: RepairMessageCreateRequest,
        files: list[UploadedFilePayload],
    ) -> RepairMessage:
        normalized_text = _normalize_message_text(payload.text)
        if normalized_text is None and not files:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Сообщение ремонта должно содержать текст или хотя бы одно вложение.",
            )

        message = RepairMessage(
            repair_id=repair.id,
            batch_key=repair.batch_key,
            author_user_id=author.id,
            author_display_name=_format_user_display_name(author),
            text=normalized_text,
        )
        self.repair_messages.add(message)
        for file_payload in files:
            self._create_repair_message_attachment(
                repair=repair,
                message=message,
                uploader=author,
                file_payload=file_payload,
        )
        return message

    def _create_verification_message_record(
        self,
        *,
        verification: Verification,
        author: User,
        payload: VerificationMessageCreateRequest,
        files: list[UploadedFilePayload],
    ) -> VerificationMessage:
        normalized_text = _normalize_message_text(payload.text)
        if normalized_text is None and not files:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Сообщение поверки должно содержать текст или хотя бы одно вложение.",
            )

        message = VerificationMessage(
            verification_id=verification.id,
            batch_key=verification.batch_key,
            author_user_id=author.id,
            author_display_name=_format_user_display_name(author),
            text=normalized_text,
        )
        self.verification_messages.add(message)
        for file_payload in files:
            self._create_verification_message_attachment(
                verification=verification,
                message=message,
                uploader=author,
                file_payload=file_payload,
            )
        return message

    def _create_repair_message_attachment(
        self,
        *,
        repair: Repair,
        message: RepairMessage,
        uploader: User,
        file_payload: UploadedFilePayload,
    ) -> RepairMessageAttachment:
        normalized_file_name = _normalize_attachment_file_name(file_payload.file_name)
        if not file_payload.content:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Attachment file must not be empty.",
            )

        storage_dir = (
            settings.attachment_storage_path
            / "repair-messages"
            / str(repair.id)
            / str(message.id)
        )
        storage_dir.mkdir(parents=True, exist_ok=True)
        storage_name = f"{uuid4().hex}{Path(normalized_file_name).suffix.lower()}"
        file_path = storage_dir / storage_name
        file_path.write_bytes(file_payload.content)

        attachment = RepairMessageAttachment(
            repair_message_id=message.id,
            uploaded_by_user_id=uploader.id,
            uploaded_by_display_name=_format_user_display_name(uploader),
            file_name=normalized_file_name,
            file_mime_type=_normalize_optional_text(file_payload.content_type),
            file_size=len(file_payload.content),
            storage_path=str(file_path.relative_to(settings.attachment_storage_path)),
        )
        self.repair_message_attachments.add(attachment)
        return attachment

    def _create_verification_message_attachment(
        self,
        *,
        verification: Verification,
        message: VerificationMessage,
        uploader: User,
        file_payload: UploadedFilePayload,
    ) -> VerificationMessageAttachment:
        normalized_file_name = _normalize_attachment_file_name(file_payload.file_name)
        if not file_payload.content:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Attachment file must not be empty.",
            )

        storage_dir = (
            settings.attachment_storage_path
            / "verification-messages"
            / str(verification.id)
            / str(message.id)
        )
        storage_dir.mkdir(parents=True, exist_ok=True)
        storage_name = f"{uuid4().hex}{Path(normalized_file_name).suffix.lower()}"
        file_path = storage_dir / storage_name
        file_path.write_bytes(file_payload.content)

        attachment = VerificationMessageAttachment(
            verification_message_id=message.id,
            uploaded_by_user_id=uploader.id,
            uploaded_by_display_name=_format_user_display_name(uploader),
            file_name=normalized_file_name,
            file_mime_type=_normalize_optional_text(file_payload.content_type),
            file_size=len(file_payload.content),
            storage_path=str(file_path.relative_to(settings.attachment_storage_path)),
        )
        self.verification_message_attachments.add(attachment)
        return attachment

    def _get_folder(self, folder_id: int) -> EquipmentFolder:
        folder = self.folders.get_by_id(folder_id)
        if folder is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Folder not found.",
            )
        return folder

    def _get_group(self, group_id: int) -> EquipmentGroup:
        group = self.groups.get_by_id(group_id)
        if group is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Group not found.",
            )
        return group

    def _get_attachment(self, *, equipment_id: int, attachment_id: int) -> EquipmentAttachment:
        self.get_equipment(equipment_id=equipment_id)
        attachment = self.attachments.get_by_id(attachment_id)
        if attachment is None or attachment.equipment_id != equipment_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Attachment not found.",
            )
        return attachment

    def _get_comment(self, *, equipment_id: int, comment_id: int) -> EquipmentComment:
        self.get_equipment(equipment_id=equipment_id)
        comment = self.comments.get_by_id(comment_id)
        if comment is None or comment.equipment_id != equipment_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Comment not found.",
            )
        return comment

    def _get_active_repair(self, *, equipment_id: int) -> Repair:
        self.get_equipment(equipment_id=equipment_id)
        repair = self.repairs.get_active_by_equipment_id(equipment_id=equipment_id)
        if repair is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Для этого прибора нет активного ремонта.",
            )
        return repair

    def _get_active_verification(self, *, equipment_id: int) -> Verification:
        equipment = self.get_equipment(equipment_id=equipment_id)
        if equipment.equipment_type != EquipmentType.SI:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Для этого прибора активная поверка недоступна.",
            )
        verification = self.verifications.get_active_by_equipment_id(
            equipment_id=equipment_id
        )
        if verification is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Для этого прибора нет активной поверки.",
            )
        return verification

    def _sync_equipment_status(self, *, equipment: Equipment) -> None:
        if self.repairs.get_active_by_equipment_id(equipment_id=equipment.id) is not None:
            equipment.status = EquipmentStatus.IN_REPAIR
            return

        if (
            equipment.equipment_type == EquipmentType.SI
            and self.verifications.get_active_by_equipment_id(equipment_id=equipment.id)
            is not None
        ):
            equipment.status = EquipmentStatus.IN_VERIFICATION
            return

        if equipment.status in {EquipmentStatus.IN_REPAIR, EquipmentStatus.IN_VERIFICATION}:
            equipment.status = EquipmentStatus.IN_WORK

    def _get_repair_message(self, *, repair: Repair, message_id: int) -> RepairMessage:
        message = self.repair_messages.get_by_id(message_id)
        if message is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Repair message not found.",
            )
        if repair.batch_key:
            if message.batch_key != repair.batch_key:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Repair message not found.",
                )
            return message
        if message.repair_id != repair.id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Repair message not found.",
            )
        return message

    def _get_verification_message(
        self,
        *,
        verification_id: int,
        message_id: int,
    ) -> VerificationMessage:
        message = self.verification_messages.get_by_id(message_id)
        if message is None or message.verification_id != verification_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Verification message not found.",
            )
        return message

    def _get_repair_message_attachment(
        self,
        *,
        message_id: int,
        attachment_id: int,
    ) -> RepairMessageAttachment:
        attachment = self.repair_message_attachments.get_by_id(attachment_id)
        if attachment is None or attachment.repair_message_id != message_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Repair attachment not found.",
            )
        return attachment

    def _get_verification_message_attachment(
        self,
        *,
        message_id: int,
        attachment_id: int,
    ) -> VerificationMessageAttachment:
        attachment = self.verification_message_attachments.get_by_id(attachment_id)
        if attachment is None or attachment.verification_message_id != message_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Verification attachment not found.",
            )
        return attachment

    def _assert_comment_owner(self, *, comment: EquipmentComment, current_user: User) -> None:
        if comment.author_user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You cannot modify this comment.",
            )

    def _assert_repair_message_owner(
        self,
        *,
        message: RepairMessage,
        current_user: User,
    ) -> None:
        if current_user.role in {UserRole.ADMINISTRATOR, UserRole.MKAIR}:
            return
        if message.author_user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You cannot delete this repair message.",
            )

    def _assert_verification_message_owner(
        self,
        *,
        message: VerificationMessage,
        current_user: User,
    ) -> None:
        if current_user.role in {UserRole.ADMINISTRATOR, UserRole.MKAIR}:
            return
        if message.author_user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You cannot delete this verification message.",
            )

    def _build_verification_queue_item(
        self,
        *,
        verification: Verification,
        equipment: Equipment,
        si_verification: SIVerification | None,
        has_active_repair: bool,
    ) -> VerificationQueueItemRead:
        return VerificationQueueItemRead(
            equipment_id=equipment.id,
            verification_id=verification.id,
            batch_key=verification.batch_key,
            batch_name=verification.batch_name,
            folder_id=equipment.folder_id,
            object_name=equipment.object_name,
            equipment_name=equipment.name,
            modification=equipment.modification,
            serial_number=equipment.serial_number,
            manufacture_year=equipment.manufacture_year,
            route_city=verification.route_city,
            route_destination=verification.route_destination,
            sent_to_verification_at=verification.sent_to_verification_at,
            received_at_destination_at=verification.received_at_destination_at,
            handed_to_csm_at=verification.handed_to_csm_at,
            verification_completed_at=verification.verification_completed_at,
            picked_up_from_csm_at=verification.picked_up_from_csm_at,
            shipped_back_at=verification.shipped_back_at,
            returned_from_verification_at=verification.returned_from_verification_at,
            closed_at=verification.closed_at,
            has_active_repair=has_active_repair,
            result_docnum=si_verification.result_docnum if si_verification else None,
            valid_date=si_verification.valid_date if si_verification else None,
            arshin_url=si_verification.arshin_url if si_verification else None,
            created_at=verification.created_at,
            updated_at=verification.updated_at,
        )

    def _build_repair_queue_item(
        self,
        *,
        repair: Repair,
        equipment: Equipment,
        si_verification: SIVerification | None,
        has_active_verification: bool,
    ) -> RepairQueueItemRead:
        registration_deadline_at = _calculate_registration_deadline_at(
            arrived_to_lensk_at=repair.arrived_to_lensk_at
        )
        control_deadline_at = _calculate_control_deadline_at(
            actually_received_at=repair.actually_received_at,
            registration_deadline_at=registration_deadline_at,
        )
        payment_deadline_at = _calculate_payment_deadline_at(
            incoming_control_at=repair.incoming_control_at,
            control_deadline_at=control_deadline_at,
        )
        repair_overdue_days = _calculate_repair_overdue_days(
            repair_deadline_at=repair.repair_deadline_at,
            arrived_to_origin_at=repair.arrived_to_lensk_at,
        )
        registration_overdue_days = _calculate_stage_overdue_days(
            deadline_at=registration_deadline_at,
            completed_at=repair.actually_received_at,
        )
        control_overdue_days = _calculate_stage_overdue_days(
            deadline_at=control_deadline_at,
            completed_at=repair.incoming_control_at,
        )
        payment_overdue_days = _calculate_stage_overdue_days(
            deadline_at=payment_deadline_at,
            completed_at=repair.paid_at,
        )
        return RepairQueueItemRead(
            repair_id=repair.id,
            equipment_id=equipment.id,
            batch_key=repair.batch_key,
            batch_name=repair.batch_name,
            folder_id=equipment.folder_id,
            object_name=equipment.object_name,
            equipment_type=equipment.equipment_type,
            equipment_name=equipment.name,
            modification=equipment.modification,
            serial_number=equipment.serial_number,
            manufacture_year=equipment.manufacture_year,
            current_location_manual=equipment.current_location_manual,
            route_city=repair.route_city,
            route_destination=repair.route_destination,
            sent_to_repair_at=repair.sent_to_repair_at,
            repair_deadline_at=repair.repair_deadline_at,
            arrived_to_destination_at=repair.arrived_to_destination_at,
            sent_from_repair_at=repair.sent_from_repair_at,
            sent_from_irkutsk_at=repair.sent_from_irkutsk_at,
            arrived_to_lensk_at=repair.arrived_to_lensk_at,
            registration_deadline_at=registration_deadline_at,
            actually_received_at=repair.actually_received_at,
            control_deadline_at=control_deadline_at,
            incoming_control_at=repair.incoming_control_at,
            payment_deadline_at=payment_deadline_at,
            paid_at=repair.paid_at,
            closed_at=repair.closed_at,
            has_active_verification=has_active_verification,
            result_docnum=si_verification.result_docnum if si_verification else None,
            current_stage_label=_build_repair_progress_label(repair=repair),
            repair_overdue_days=repair_overdue_days,
            registration_overdue_days=registration_overdue_days,
            control_overdue_days=control_overdue_days,
            payment_overdue_days=payment_overdue_days,
            max_overdue_days=max(
                repair_overdue_days,
                registration_overdue_days,
                control_overdue_days,
                payment_overdue_days,
            ),
            created_at=repair.created_at,
            updated_at=repair.updated_at,
        )

    def _build_existing_si_message(
        self,
        existing: SIVerification,
        *,
        prefix: str,
    ) -> str:
        equipment = existing.equipment
        if equipment is None:
            return f"{prefix} к другому прибору."

        folder_name: str | None = None
        if equipment.folder_id is not None:
            folder = self.folders.get_by_id(equipment.folder_id)
            if folder is not None:
                folder_name = folder.name

        message = f"{prefix} за прибором «{equipment.name}»"
        if folder_name:
            message = f"{message} в папке «{folder_name}»"
        return f"{message}."

    def _validate_si_payload_for_create(self, payload: EquipmentCreateRequest) -> None:
        if payload.equipment_type == EquipmentType.SI:
            if payload.si_verification is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                    detail="SI equipment must be created from an Arshin search result.",
                )
            if payload.si_verification.detail_payload_json is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                    detail="SI equipment creation requires Arshin detail fetch by vri_id.",
                )

            existing = self.si_verifications.get_by_vri_id(
                vri_id=payload.si_verification.vri_id.strip()
            )
            if existing is not None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=self._build_existing_si_message(
                        existing,
                        prefix="Прибор с этой записью Аршина уже существует",
                    ),
                )
            return

        if payload.si_verification is not None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="SI verification data is allowed only for SI equipment.",
            )


def _build_repair_progress_label(*, repair: Repair) -> str:
    if repair.closed_at is not None:
        return "Ремонт завершен"
    if repair.paid_at is not None:
        return "Оплата выполнена"
    if repair.incoming_control_at is not None:
        return "Ожидает оплаты"
    if repair.actually_received_at is not None:
        return "На входном контроле"
    if repair.arrived_to_lensk_at is not None:
        return "Ожидает получения"
    if repair.sent_from_irkutsk_at is not None:
        return "В пути обратно"
    if repair.sent_from_repair_at is not None:
        return "Ремонт произведен"
    if repair.arrived_to_destination_at is not None:
        return "В ремонте"
    return "В ремонте"


def _get_verification_progress_label(item: VerificationQueueItemRead) -> str:
    if item.closed_at is not None:
        return "Поверка завершена"
    if item.returned_from_verification_at is not None:
        return "Получено обратно"
    if item.shipped_back_at is not None:
        return "Ожидает получения"
    if item.picked_up_from_csm_at is not None:
        return "Получено в ЦСМ"
    if item.verification_completed_at is not None:
        return "Поверка выполнена"
    if item.handed_to_csm_at is not None:
        return "В ЦСМ на поверке"
    if item.received_at_destination_at is not None:
        return "Получено в пункте назначения"
    return "Отправлено в поверку"


def _build_repair_archive_name(
    *,
    folder_name: str,
    start_date: date,
    end_date: date | None,
) -> str:
    normalized_folder_name = re.sub(r'[\\/:*?"<>|]+', " ", folder_name).strip()
    normalized_folder_name = re.sub(r"\s+", " ", normalized_folder_name) or "Без папки"
    end_label = end_date.strftime("%d.%m.%Y") if end_date else "без даты"
    return (
        f"Ремонт {normalized_folder_name} "
        f"{start_date.strftime('%d.%m.%Y')} по {end_label}"
    )


def _build_verification_archive_name(
    *,
    label: str,
    start_date: date,
    end_date: date | None,
) -> str:
    normalized_label = re.sub(r'[\\/:*?"<>|]+', " ", label).strip()
    normalized_label = re.sub(r"\s+", " ", normalized_label) or "Без папки"
    end_label = end_date.strftime("%d.%m.%Y") if end_date else "без даты"
    return (
        f"Поверка {normalized_label} "
        f"{start_date.strftime('%d.%m.%Y')} по {end_label}"
    )


def _build_equipment_batch_member_label(equipment: Equipment) -> str:
    serial_number = equipment.serial_number
    if not serial_number and equipment.si_verification is not None:
        serial_number = equipment.si_verification.mi_number
    if serial_number:
        return f"{equipment.name} (зав. № {serial_number})"
    return equipment.name


def _calculate_repair_overdue_days(
    *,
    repair_deadline_at: date,
    arrived_to_origin_at: date | None,
) -> int:
    comparison_date = arrived_to_origin_at or date.today()
    return max((comparison_date - repair_deadline_at).days, 0)


def _calculate_registration_deadline_at(*, arrived_to_lensk_at: date | None) -> date | None:
    if arrived_to_lensk_at is None:
        return None
    return arrived_to_lensk_at + timedelta(days=5)


def _calculate_control_deadline_at(
    *,
    actually_received_at: date | None,
    registration_deadline_at: date | None,
) -> date | None:
    anchor = actually_received_at or registration_deadline_at
    if anchor is None:
        return None
    return anchor + timedelta(days=40)


def _calculate_payment_deadline_at(
    *,
    incoming_control_at: date | None,
    control_deadline_at: date | None,
) -> date | None:
    anchor = incoming_control_at or control_deadline_at
    if anchor is None:
        return None
    return anchor + timedelta(days=70)


def _calculate_stage_overdue_days(
    *,
    deadline_at: date | None,
    completed_at: date | None,
) -> int:
    if deadline_at is None:
        return 0
    comparison_date = completed_at or date.today()
    return max((comparison_date - deadline_at).days, 0)


def _validate_repair_milestone_order(
    *,
    route_city: str,
    route_destination: str,
    sent_to_repair_at: date,
    arrived_to_destination_at: date | None,
    sent_from_repair_at: date | None,
    sent_from_irkutsk_at: date | None,
    arrived_to_lensk_at: date | None,
    actually_received_at: date | None,
    incoming_control_at: date | None,
    paid_at: date | None,
) -> None:
    _validate_milestone_order(
        milestones=(
            (f"Отправлено в {route_destination}", sent_to_repair_at),
            (f"Прибыло в {route_destination}", arrived_to_destination_at),
            ("Ремонт произведен", sent_from_repair_at),
            (f"Отправлено в {route_city}", sent_from_irkutsk_at),
            (f"Прибыло в {route_city}", arrived_to_lensk_at),
            (f"Получено в {route_city}", actually_received_at),
            ("Входной контроль выполнен", incoming_control_at),
            ("Оплата выполнена", paid_at),
        )
    )


def _validate_verification_milestone_order(
    *,
    route_destination: str,
    sent_to_verification_at: date,
    received_at_destination_at: date | None,
    handed_to_csm_at: date | None,
    verification_completed_at: date | None,
    picked_up_from_csm_at: date | None,
    shipped_back_at: date | None,
    returned_from_verification_at: date | None,
) -> None:
    _validate_milestone_order(
        milestones=(
            ("Отправлено в поверку", sent_to_verification_at),
            (f"Получение в {route_destination}", received_at_destination_at),
            ("Передано в ЦСМ", handed_to_csm_at),
            ("Поверка выполнена", verification_completed_at),
            ("Получено в ЦСМ", picked_up_from_csm_at),
            ("Упаковано и отправлено обратно", shipped_back_at),
            ("Получено обратно", returned_from_verification_at),
        )
    )


def _validate_milestone_order(
    *, milestones: tuple[tuple[str, date | None], ...]
) -> None:
    missing_previous_label: str | None = None
    previous_label: str | None = None
    previous_value: date | None = None

    for label, value in milestones:
        if value is None:
            if missing_previous_label is None:
                missing_previous_label = label
            continue

        if missing_previous_label is not None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail=(
                    f"Этап «{label}» нельзя указать раньше, чем этап "
                    f"«{missing_previous_label}»."
                ),
            )

        if previous_value is not None and value < previous_value:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail=(
                    f"Этап «{label}» не может быть раньше этапа "
                    f"«{previous_label}»."
                ),
            )

        previous_label = label
        previous_value = value


def _normalize_required_text(value: str | None, *, field_label: str) -> str:
    if value is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"{field_label} must not be empty.",
        )
    normalized = value.strip()
    if not normalized:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"{field_label} must not be empty.",
        )
    if len(normalized) > 255:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"{field_label} is too long. Maximum length is 255 characters.",
        )
    return normalized


def _normalize_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    if not normalized:
        return None
    if len(normalized) > 255:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Field is too long. Maximum length is 255 characters.",
        )
    return normalized


def _normalize_long_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    if not normalized:
        return None
    if len(normalized) > 1024:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Field is too long. Maximum length is 1024 characters.",
        )
    return normalized


def _normalize_attachment_file_name(value: str | None) -> str:
    candidate = Path(value or "").name.strip()
    if not candidate:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Attachment file name must not be empty.",
        )
    if len(candidate) > 255:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Attachment file name is too long. Maximum length is 255 characters.",
        )
    return candidate


def _normalize_comment_text(value: str | None) -> str:
    if value is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Comment text must not be empty.",
        )
    normalized = value.strip()
    if not normalized:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Comment text must not be empty.",
        )
    if len(normalized) > 4000:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Comment text is too long. Maximum length is 4000 characters.",
        )
    return normalized


def _normalize_message_text(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    if not normalized:
        return None
    if len(normalized) > 4000:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Message text is too long. Maximum length is 4000 characters.",
        )
    return normalized


def _validate_manufacture_year(value: int | None) -> int | None:
    if value is None:
        return None
    if value < 1900 or value > 2100:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Manufacture year must be between 1900 and 2100.",
        )
    return value


def _format_user_display_name(user: User) -> str:
    parts = [user.last_name.strip(), user.first_name.strip()]
    if user.patronymic and user.patronymic.strip():
        parts.append(user.patronymic.strip())
    return " ".join(part for part in parts if part)


def _extract_si_detail_single_mi(detail: dict | None) -> dict:
    if not isinstance(detail, dict):
        return {}
    mi_info = detail.get("miInfo")
    if not isinstance(mi_info, dict):
        return {}
    for key in ("singleMI", "mi", "etaMI"):
        value = mi_info.get(key)
        if isinstance(value, dict):
            return value
    return {}


def _extract_si_detail_vri_info(detail: dict | None) -> dict:
    if not isinstance(detail, dict):
        return {}
    vri_info = detail.get("vriInfo")
    return vri_info if isinstance(vri_info, dict) else {}


def _extract_cert_num_from_detail(vri_info: dict) -> str | None:
    applicable = vri_info.get("applicable")
    if isinstance(applicable, dict):
        return _first_nonempty_str(applicable.get("certNum"), applicable.get("certificateNumber"))
    return _first_nonempty_str(vri_info.get("certNum"))


def _first_nonempty_str(*values: object) -> str | None:
    for value in values:
        if value is None:
            continue
        normalized = str(value).strip()
        if normalized:
            return normalized
    return None


def _first_nonempty_int(*values: object) -> int | None:
    for value in values:
        if value is None or value == "":
            continue
        try:
            return int(value)
        except (TypeError, ValueError):
            continue
    return None


def _parse_date_to_datetime(value: object) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if not isinstance(value, str):
        return None

    candidate = value.strip()
    if not candidate:
        return None
    candidate = candidate.replace("Z", "+00:00")

    try:
        return datetime.fromisoformat(candidate)
    except ValueError:
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%d.%m.%Y %H:%M:%S"):
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


def _first_nonempty_datetime(*values: datetime | None) -> datetime | None:
    for value in values:
        if value is not None:
            return value
    return None


def _build_si_create_request_from_arshin(
    *,
    folder_id: int,
    object_name: str,
    status_value: EquipmentStatus,
    current_location_manual: str | None,
    result: ArshinSearchResultRead,
    detail: ArshinVriDetailRead,
) -> EquipmentCreateRequest:
    return EquipmentCreateRequest(
        folder_id=folder_id,
        group_id=None,
        object_name=object_name,
        equipment_type=EquipmentType.SI,
        name=detail.type_name or result.mit_title or "СИ из Аршина",
        modification=detail.modification or result.mi_modification,
        serial_number=detail.serial_number or result.mi_number,
        manufacture_year=detail.manufacture_year,
        status=status_value,
        current_location_manual=current_location_manual,
        si_verification={
            "vri_id": result.vri_id,
            "arshin_url": result.arshin_url,
            "org_title": result.org_title,
            "mit_number": result.mit_number,
            "mit_title": result.mit_title,
            "mit_notation": result.mit_notation,
            "mi_number": result.mi_number,
            "result_docnum": result.result_docnum,
            "verification_date": result.verification_date,
            "valid_date": result.valid_date,
            "raw_payload_json": result.raw_payload_json,
            "detail_payload_json": detail.raw_payload_json,
        },
    )


def _select_bulk_import_candidate(
    *,
    certificate_number: str,
    results: list[ArshinSearchResultRead],
) -> ArshinSearchResultRead | None:
    if not results:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Arshin did not return any records for this certificate number.",
        )

    if len(results) == 1:
        return results[0]

    normalized_input = _normalize_certificate_number(certificate_number)
    exact_matches = [
        result
        for result in results
        if _normalize_certificate_number(result.result_docnum) == normalized_input
    ]
    if len(exact_matches) == 1:
        return exact_matches[0]

    return None


def _normalize_certificate_number(value: str | None) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", value).upper()


@dataclass(slots=True)
class ParsedCertificateImportRow:
    row_number: int
    certificate_number: str
    verification_year: int | None = None


@dataclass(slots=True)
class UploadedFilePayload:
    file_name: str | None
    content_type: str | None
    content: bytes


_VERIFICATION_MILESTONE_LABELS: tuple[tuple[str, str], ...] = (
    ("received_at_destination_at", "Получено в пункте назначения"),
    ("handed_to_csm_at", "Передано в ЦСМ"),
    ("verification_completed_at", "Поверка выполнена"),
    ("picked_up_from_csm_at", "Получено в ЦСМ"),
    ("shipped_back_at", "Упаковано и отправлено обратно"),
    ("returned_from_verification_at", "Получено обратно"),
)

_REPAIR_MILESTONE_LABELS: tuple[tuple[str, str], ...] = (
    ("arrived_to_destination_at", "Прибыло в пункт назначения"),
    ("sent_from_repair_at", "Ремонт произведен"),
    ("sent_from_irkutsk_at", "Отправлено в пункт отправления"),
    ("arrived_to_lensk_at", "Прибыло в пункт отправления"),
    ("actually_received_at", "Получено"),
    ("incoming_control_at", "Входной контроль выполнен"),
    ("paid_at", "Оплата выполнена"),
)


def _collect_changed_repair_milestone_labels(
    *,
    payload: RepairMilestonesUpdateRequest,
    route_destination: str,
) -> list[str]:
    labels: list[str] = []
    if "sent_to_repair_at" in payload.model_fields_set:
        labels.append(f"Отправлено в {route_destination}")
    for field_name, milestone_label in _REPAIR_MILESTONE_LABELS:
        if field_name in payload.model_fields_set:
            labels.append(milestone_label)
    return labels


def _collect_changed_verification_milestone_labels(
    *,
    payload: VerificationMilestonesUpdateRequest,
    route_destination: str,
) -> list[str]:
    labels: list[str] = []
    for field_name, milestone_label in _VERIFICATION_MILESTONE_LABELS:
        if field_name not in payload.model_fields_set:
            continue
        if field_name == "received_at_destination_at":
            labels.append(f"Получение в {route_destination}")
        else:
            labels.append(milestone_label)
    return labels


def _build_preview_description(text: str | None) -> str | None:
    normalized = _normalize_message_text(text)
    if normalized is None:
        return None
    if len(normalized) <= 160:
        return normalized
    return normalized[:157].rstrip() + "..."


def _build_message_event_description(
    *,
    text: str | None,
    files: list[UploadedFilePayload],
) -> str | None:
    parts: list[str] = []
    preview = _build_preview_description(text)
    if preview:
        parts.append(preview)
    if files:
        parts.append(f"Вложений: {len(files)}.")
    return " ".join(parts) if parts else None


def _build_named_detail(label: str, value: object | None) -> str | None:
    if value is None:
        return None
    rendered = str(value).strip()
    if not rendered:
        return None
    return f"{label}: {rendered}"


def _build_nonempty_description(parts: list[str | None]) -> str | None:
    filtered = [part for part in parts if part]
    return " • ".join(filtered) if filtered else None


def _build_verification_milestone_message(
    *,
    current_user: User,
    milestone_label: str,
    milestone_date,
    route_destination: str,
) -> str:
    resolved_label = milestone_label
    if milestone_label == "Получено в пункте назначения":
        resolved_label = f"Получено в пункте назначения ({route_destination})"
    return (
        f"{_format_user_display_name(current_user)} отметил этап "
        f"«{resolved_label}» ({_format_short_date(milestone_date)})."
    )


def _build_repair_milestone_message(
    *,
    current_user: User,
    milestone_label: str,
    milestone_date,
) -> str:
    return (
        f"{_format_user_display_name(current_user)} отметил этап "
        f"«{milestone_label}» ({_format_short_date(milestone_date)})."
    )


def _format_short_date(value) -> str:
    return value.strftime("%d.%m.%Y")


def _extract_certificate_rows_from_table(
    *,
    file_name: str | None,
    content: bytes,
) -> list[ParsedCertificateImportRow]:
    lower_name = (file_name or "").lower()
    if lower_name.endswith(".csv"):
        return _extract_certificate_rows_from_csv(content)
    return _extract_certificate_rows_from_workbook(content)


def _extract_certificate_rows_from_csv(content: bytes) -> list[ParsedCertificateImportRow]:
    decoded = content.decode("utf-8-sig")
    reader = csv.reader(StringIO(decoded))
    rows = list(reader)
    return _extract_certificate_rows_from_matrix(rows)


def _extract_certificate_rows_from_workbook(content: bytes) -> list[ParsedCertificateImportRow]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        matrix: list[list[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            matrix.append(["" if cell is None else str(cell).strip() for cell in row])
        return _extract_certificate_rows_from_matrix(matrix)
    finally:
        workbook.close()


def _extract_certificate_rows_from_matrix(
    rows: list[list[str]],
) -> list[ParsedCertificateImportRow]:
    if not rows:
        return []

    header_row_index, cert_column, date_column = _detect_certificate_layout(rows)
    start_index = header_row_index + 1 if header_row_index is not None else 0
    if cert_column is None:
        cert_column = 0

    extracted: list[ParsedCertificateImportRow] = []
    for index, row in enumerate(rows[start_index:], start=start_index + 1):
        if cert_column >= len(row):
            continue
        certificate_number = row[cert_column].strip()
        if not certificate_number:
            continue
        verification_year = None
        if date_column is not None and date_column < len(row):
            verification_year = _extract_year_from_cell(row[date_column])
        extracted.append(
            ParsedCertificateImportRow(
                row_number=index,
                certificate_number=certificate_number,
                verification_year=verification_year,
            )
        )
    return extracted


def _detect_certificate_layout(
    rows: list[list[str]],
) -> tuple[int | None, int | None, int | None]:
    max_header_scan_rows = min(len(rows), 20)
    for row_index in range(max_header_scan_rows):
        cert_column = _detect_certificate_column(rows[row_index])
        if cert_column is not None:
            date_column = _detect_verification_date_column(rows[row_index])
            return row_index, cert_column, date_column
    return None, None, None


def _detect_certificate_column(header_row: list[str]) -> int | None:
    for index, value in enumerate(header_row):
        normalized = _normalize_header_label(value)
        if not normalized:
            continue
        if any(
            keyword in normalized
            for keyword in ("свид", "certificate", "документ", "document")
        ):
            return index
    return None


def _detect_verification_date_column(header_row: list[str]) -> int | None:
    for index, value in enumerate(header_row):
        normalized = _normalize_header_label(value)
        if not normalized:
            continue
        if normalized == "дата поверки" or normalized == "verification date":
            return index
        if "дата" in normalized and "поверк" in normalized:
            return index
        if "verification" in normalized and "date" in normalized:
            return index
    return None


def _normalize_header_label(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def _extract_year_from_cell(value: str) -> int | None:
    match = re.search(r"(19|20)\d{2}", value)
    if not match:
        return None
    year = int(match.group(0))
    if 1900 <= year <= 2100:
        return year
    return None
